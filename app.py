from flask import Flask, request, jsonify, send_file, Response, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
import os, io, sqlite3, json, shutil, threading, time, socket, re

app = Flask(__name__)
app.secret_key = b'\x9f\x4a\x2c\x8e\x1b\x7d\x3f\x56\xa0\xc2\x4e\x91\x68\x0b\xd5\x37'
app.permanent_session_lifetime = timedelta(hours=8)

DB_FILE    = "passports.db"
BACKUP_DIR = "backups"
MAX_BACKUPS = 7

# ── DB ───────────────────────────────────────────────────────

def get_db():
    db = sqlite3.connect(DB_FILE, check_same_thread=False)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    return db

DEFAULT_PREFIXES = {
    "fagh":"A", "amsa":"B", "cuba":"C",
    "safar":"D", "top":"E", "top10":"F"
}

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS groups (
            id     TEXT PRIMARY KEY,
            name   TEXT NOT NULL,
            prefix TEXT
        );
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role          TEXT NOT NULL DEFAULT 'worker',
            group_id      TEXT,
            active        INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS pilgrims (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            passport       TEXT UNIQUE NOT NULL,
            group_id       TEXT NOT NULL,
            seq            INTEGER,
            seq_code       TEXT,
            status         TEXT DEFAULT '',
            is_diplomatic  INTEGER DEFAULT 0,
            departed       TEXT DEFAULT 'لا',
            departed_at    TEXT,
            notes          TEXT DEFAULT '',
            created_at     TEXT,
            updated_at     TEXT
        );
        CREATE TABLE IF NOT EXISTS audit_logs (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            username   TEXT,
            user_group TEXT,
            action     TEXT,
            passport   TEXT,
            old_values TEXT,
            new_values TEXT,
            ts         TEXT
        );
    """)
    # safe ALTER for existing databases
    for stmt in ["ALTER TABLE groups   ADD COLUMN prefix        TEXT",
                 "ALTER TABLE pilgrims ADD COLUMN seq_code     TEXT",
                 "ALTER TABLE pilgrims ADD COLUMN is_diplomatic INTEGER DEFAULT 0"]:
        try: db.execute(stmt)
        except: pass
    for gid, gname in [("fagh","فاغبين"),("amsa","امسا"),("cuba","كوبا"),
                        ("safar","سفريات العمرة"),("top","تروبك"),("top10","توب ١٠")]:
        db.execute("INSERT OR IGNORE INTO groups (id,name,prefix) VALUES (?,?,?)",
                   (gid, gname, DEFAULT_PREFIXES.get(gid,"")))
    # set missing prefixes on existing rows
    for gid, prefix in DEFAULT_PREFIXES.items():
        db.execute("UPDATE groups SET prefix=? WHERE id=? AND (prefix IS NULL OR prefix='')",
                   (prefix, gid))
    if not db.execute("SELECT 1 FROM users WHERE username='admin'").fetchone():
        db.execute("INSERT INTO users (username,password_hash,role) VALUES (?,?,?)",
                   ("admin", generate_password_hash("admin2024"), "admin"))
    for gid in DEFAULT_PREFIXES:
        if not db.execute("SELECT 1 FROM users WHERE username=?", (gid,)).fetchone():
            db.execute("INSERT INTO users (username,password_hash,role,group_id) VALUES (?,?,?,?)",
                       (gid, generate_password_hash("1234"), "worker", gid))
    db.commit(); db.close()

def _group_prefix(db, group_id):
    row = db.execute("SELECT prefix FROM groups WHERE id=?", (group_id,)).fetchone()
    if row and row["prefix"]:
        return row["prefix"]
    return DEFAULT_PREFIXES.get(group_id, group_id[0].upper() if group_id else "X")

def next_seq_code(db, group_id):
    """Return (seq_int, seq_code) for next record in group — never reuses after delete."""
    max_seq = db.execute(
        "SELECT COALESCE(MAX(seq),0) as m FROM pilgrims WHERE group_id=?", (group_id,)
    ).fetchone()["m"]
    seq = max_seq + 1
    prefix = _group_prefix(db, group_id)
    return seq, f"{prefix}-{seq}"

def migrate_seq_codes():
    """Backfill seq_code for existing records that don't have one."""
    db = get_db()
    rows = db.execute("""
        SELECT p.id, p.seq, p.group_id
        FROM pilgrims p
        WHERE p.seq_code IS NULL OR p.seq_code = ''
    """).fetchall()
    for r in rows:
        prefix = _group_prefix(db, r["group_id"])
        code   = f"{prefix}-{r['seq'] or 0}"
        db.execute("UPDATE pilgrims SET seq_code=? WHERE id=?", (code, r["id"]))
    if rows:
        db.commit()
        print(f"✅ تم إنشاء seq_code لـ {len(rows)} سجل")
    db.close()

def migrate_json():
    if not os.path.exists("data.json") or os.path.exists("data.json.migrated"):
        return
    try:
        with open("data.json","r",encoding="utf-8") as f:
            data = json.load(f)
        db = get_db()
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        for gid, rows in data.items():
            for r in rows:
                db.execute("""INSERT OR IGNORE INTO pilgrims
                    (passport,group_id,seq,status,departed,notes,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (r.get("passport","").upper(), gid, r.get("seq"),
                     r.get("status",""), r.get("departed","لا"),
                     r.get("notes",""), r.get("updated",now), r.get("updated",now)))
        db.commit(); db.close()
        open("data.json.migrated","w").close()
        print("✅ تم نقل data.json إلى SQLite")
    except Exception as e:
        print(f"⚠️ خطأ في النقل: {e}")

# ── Helpers ──────────────────────────────────────────────────

def log_action(action, passport=None, old=None, new=None):
    try:
        db = get_db()
        db.execute("INSERT INTO audit_logs (username,user_group,action,passport,old_values,new_values,ts) VALUES (?,?,?,?,?,?,?)",
            (session.get("username","?"), session.get("group_id",""), action, passport,
             json.dumps(old, ensure_ascii=False) if old else None,
             json.dumps(new, ensure_ascii=False) if new else None,
             datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        db.commit(); db.close()
    except: pass

def require_login(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("user_id"):
            return jsonify({"ok":False,"msg":"غير مسجّل الدخول","unauth":True}), 401
        return f(*a,**k)
    return d

def require_admin(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("user_id") or session.get("role") != "admin":
            return jsonify({"ok":False,"msg":"غير مصرح"}), 403
        return f(*a,**k)
    return d

def do_backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_DIR, f"passports_{ts}.db")
    shutil.copy2(DB_FILE, dst)
    files = sorted(f for f in os.listdir(BACKUP_DIR) if f.endswith(".db"))
    for old in files[:-MAX_BACKUPS]:
        os.remove(os.path.join(BACKUP_DIR, old))
    return dst

def backup_scheduler():
    while True:
        time.sleep(6*3600)
        try: do_backup()
        except: pass

# ── Auth ─────────────────────────────────────────────────────

@app.route("/api/public/groups")
def public_groups():
    db = get_db()
    groups = db.execute("SELECT id,name FROM groups ORDER BY name").fetchall()
    db.close()
    return jsonify({"ok":True,"groups":[dict(g) for g in groups]})

@app.route("/api/login", methods=["POST"])
def login():
    d = request.json or {}
    username = d.get("username","").strip()
    pw       = d.get("pw","")
    gid      = d.get("gid","")          # "admin" or a group id
    if not username or not pw:
        return jsonify({"ok":False,"msg":"أدخل اسم المستخدم وكلمة السر"})
    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE username=? AND active=1",(username,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], pw):
        db.close()
        return jsonify({"ok":False,"msg":"اسم المستخدم أو كلمة السر غلط"})
    if gid == "admin":
        if user["role"] != "admin":
            db.close()
            return jsonify({"ok":False,"msg":"هذا المستخدم ليس مشرفاً"})
    else:
        if user["role"] == "admin":
            db.close()
            return jsonify({"ok":False,"msg":"استخدم خيار المشرف للدخول"})
        if user["group_id"] != gid:
            db.close()
            return jsonify({"ok":False,"msg":"هذا المستخدم لا ينتمي لهذه المجموعة"})
    grp = db.execute("SELECT name FROM groups WHERE id=?", (user["group_id"],)).fetchone() if user["group_id"] else None
    db.close()
    session.clear()
    session.permanent = True
    session["user_id"]  = user["id"]
    session["username"] = user["username"]
    session["role"]     = user["role"]
    session["group_id"] = user["group_id"] or ""
    session["name"]     = "مشرف عام" if user["role"]=="admin" else (grp["name"] if grp else username)
    log_action("دخول")
    return jsonify({"ok":True,"role":user["role"],"name":session["name"],"gid":session["group_id"]})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    log_action("خروج")
    session.clear()
    return jsonify({"ok":True})

@app.route("/api/session")
def check_session():
    if session.get("user_id"):
        return jsonify({"ok":True,"role":session["role"],"name":session["name"],"gid":session["group_id"]})
    return jsonify({"ok":False})

# ── Stats ────────────────────────────────────────────────────

@app.route("/api/stats")
@require_login
def stats():
    db  = get_db()
    gid = request.args.get("gid","")
    if session["role"] != "admin":
        gid = session["group_id"]
    if gid and gid != "all":
        rows = db.execute("SELECT status,departed FROM pilgrims WHERE group_id=?",(gid,)).fetchall()
        db.close()
        return jsonify({"ok":True,"total":len(rows),
            "active":  sum(1 for r in rows if r["status"]=="مفعل"),
            "inactive":sum(1 for r in rows if r["status"]=="غير مفعل"),
            "departed":sum(1 for r in rows if r["departed"]=="نعم")})
    groups = db.execute("SELECT * FROM groups").fetchall()
    result=[]; total=active=inactive=departed=0
    for g in groups:
        rows = db.execute("SELECT status,departed FROM pilgrims WHERE group_id=?",(g["id"],)).fetchall()
        t=len(rows); a=sum(1 for r in rows if r["status"]=="مفعل")
        i=sum(1 for r in rows if r["status"]=="غير مفعل")
        dp=sum(1 for r in rows if r["departed"]=="نعم")
        total+=t; active+=a; inactive+=i; departed+=dp
        result.append({"id":g["id"],"name":g["name"],"total":t,"active":a,"inactive":i,"departed":dp})
    db.close()
    return jsonify({"ok":True,"total":total,"active":active,"inactive":inactive,
                    "departed":departed,"groups":result})

# ── Search ───────────────────────────────────────────────────

@app.route("/api/search")
@require_login
def search():
    passport = request.args.get("passport","").strip().upper()
    if not passport:
        return jsonify({"ok":False,"msg":"أدخل رقم الجواز"})
    db = get_db()
    if session["role"] == "admin":
        row = db.execute("""SELECT p.*,g.name as group_name FROM pilgrims p
            JOIN groups g ON p.group_id=g.id WHERE UPPER(p.passport)=?""",(passport,)).fetchone()
        db.close()
        if not row: return jsonify({"ok":False,"msg":"غير موجود في أي مجموعة"})
        return jsonify({"ok":True,**dict(row)})
    row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=? AND group_id=?",
                     (passport,session["group_id"])).fetchone()
    db.close()
    if not row: return jsonify({"ok":False,"msg":"رقم الجواز غير موجود في مجموعتك"})
    return jsonify({"ok":True,**dict(row)})

# ── Add ──────────────────────────────────────────────────────

@app.route("/api/add", methods=["POST"])
@require_login
def add():
    d  = request.json or {}
    gid = session["group_id"] if session["role"]!="admin" else d.get("gid","")
    if not gid: return jsonify({"ok":False,"msg":"حدد المجموعة"})
    passport = d.get("passport","").strip().upper()
    if not passport: return jsonify({"ok":False,"msg":"أدخل رقم الجواز"})
    if not re.match(r'^[A-Z]{2}[0-9]{7}$', passport):
        return jsonify({"ok":False,"msg":"صيغة الجواز غير صحيحة — يجب حرفان + 7 أرقام (مثال: AB1234567)"})
    db = get_db()
    ex = db.execute("SELECT group_id FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
    if ex:
        grp = db.execute("SELECT name FROM groups WHERE id=?",(ex["group_id"],)).fetchone()
        db.close()
        return jsonify({"ok":False,"msg":f"⚠️ مكرر — موجود في {grp['name'] if grp else ex['group_id']}"})
    seq, seq_code = next_seq_code(db, gid)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    status = d.get("status",""); notes = d.get("notes","")
    is_diplomatic = 1 if passport.startswith("DA") else 0
    db.execute("""INSERT INTO pilgrims
        (passport,group_id,seq,seq_code,status,is_diplomatic,departed,notes,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?)""",
               (passport,gid,seq,seq_code,status,is_diplomatic,"لا",notes,now,now))
    db.commit(); db.close()
    log_action("إضافة", passport, new={"status":status,"seq_code":seq_code,"is_diplomatic":is_diplomatic})
    return jsonify({"ok":True,"seq":seq,"seq_code":seq_code,"is_diplomatic":is_diplomatic})

# ── Update ───────────────────────────────────────────────────

@app.route("/api/update", methods=["POST"])
@require_login
def update():
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    db = get_db()
    row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
    if not row: db.close(); return jsonify({"ok":False,"msg":"السجل غير موجود"})
    if session["role"]!="admin" and row["group_id"]!=session["group_id"]:
        db.close(); return jsonify({"ok":False,"msg":"غير مصرح"})
    old  = dict(row)
    nst  = d.get("status",  row["status"])
    nnt  = d.get("notes",   row["notes"] or "")
    now  = datetime.now().strftime("%Y-%m-%d %H:%M")
    if session["role"] == "admin" and "is_diplomatic" in d:
        ndip = 1 if d["is_diplomatic"] else 0
        db.execute("UPDATE pilgrims SET status=?,notes=?,is_diplomatic=?,updated_at=? WHERE id=?",
                   (nst,nnt,ndip,now,row["id"]))
    else:
        ndip = row["is_diplomatic"] if "is_diplomatic" in row.keys() else 0
        db.execute("UPDATE pilgrims SET status=?,notes=?,updated_at=? WHERE id=?",(nst,nnt,now,row["id"]))
    db.commit(); db.close()
    log_action("تعديل", passport,
               old={"status":old["status"],"notes":old["notes"],"is_diplomatic":old.get("is_diplomatic",0)},
               new={"status":nst,"notes":nnt,"is_diplomatic":ndip})
    return jsonify({"ok":True})

# ── Depart ───────────────────────────────────────────────────

@app.route("/api/depart", methods=["POST"])
@require_login
def depart():
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    db = get_db()
    row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
    if not row: db.close(); return jsonify({"ok":False,"msg":"رقم الجواز غير موجود"})
    if session["role"]!="admin" and row["group_id"]!=session["group_id"]:
        db.close(); return jsonify({"ok":False,"msg":"غير مصرح"})
    if row["departed"]=="نعم":
        db.close(); return jsonify({"ok":False,"msg":f"غادر مسبقاً في {row['departed_at']}"})
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    db.execute("UPDATE pilgrims SET departed='نعم',departed_at=?,updated_at=? WHERE id=?",(now,now,row["id"]))
    db.commit(); db.close()
    log_action("مغادرة", passport, old={"departed":"لا"}, new={"departed":"نعم","departed_at":now})
    return jsonify({"ok":True,"departed_at":now})

# ── List ─────────────────────────────────────────────────────

@app.route("/api/list")
@require_login
def list_records():
    gid  = request.args.get("gid","")
    f    = request.args.get("filter","all")
    page = max(1, int(request.args.get("page",1)))
    per  = min(200, int(request.args.get("per",50)))
    if session["role"]!="admin":
        gid = session["group_id"]
    cond=[]; params=[]
    if gid: cond.append("group_id=?"); params.append(gid)
    if f=="active":    cond.append("status='مفعل'")
    elif f=="inactive":cond.append("status='غير مفعل'")
    elif f=="departed":cond.append("departed='نعم'")
    elif f=="pending": cond.append("(departed IS NULL OR departed!='نعم')")
    where = ("WHERE "+" AND ".join(cond)) if cond else ""
    db    = get_db()
    total = db.execute(f"SELECT COUNT(*) as c FROM pilgrims {where}",params).fetchone()["c"]
    rows  = db.execute(f"SELECT * FROM pilgrims {where} ORDER BY seq LIMIT ? OFFSET ?",
                       params+[per,(page-1)*per]).fetchall()
    db.close()
    return jsonify({"ok":True,"rows":[dict(r) for r in rows],"total":total,"page":page,"per":per})

# ── Delete ───────────────────────────────────────────────────

@app.route("/api/delete", methods=["POST"])
@require_login
def delete():
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    db = get_db()
    row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
    if not row: db.close(); return jsonify({"ok":False,"msg":"السجل غير موجود"})
    if session["role"]!="admin" and row["group_id"]!=session["group_id"]:
        db.close(); return jsonify({"ok":False,"msg":"غير مصرح"})
    old = dict(row)
    db.execute("DELETE FROM pilgrims WHERE id=?",(row["id"],))
    db.commit(); db.close()
    log_action("حذف", passport, old=old)
    return jsonify({"ok":True})

# ── Audit Logs ───────────────────────────────────────────────

@app.route("/api/logs")
@require_admin
def get_logs():
    page = max(1, int(request.args.get("page",1)))
    per  = min(200, int(request.args.get("per",50)))
    cond=[]; params=[]
    u_f = request.args.get("user","")
    a_f = request.args.get("action","")
    p_f = request.args.get("passport","").upper()
    if u_f: cond.append("username=?"); params.append(u_f)
    if a_f: cond.append("action=?");   params.append(a_f)
    if p_f: cond.append("UPPER(passport) LIKE ?"); params.append(f"%{p_f}%")
    where = ("WHERE "+" AND ".join(cond)) if cond else ""
    db    = get_db()
    total = db.execute(f"SELECT COUNT(*) as c FROM audit_logs {where}",params).fetchone()["c"]
    rows  = db.execute(f"SELECT * FROM audit_logs {where} ORDER BY id DESC LIMIT ? OFFSET ?",
                       params+[per,(page-1)*per]).fetchall()
    db.close()
    return jsonify({"ok":True,"rows":[dict(r) for r in rows],"total":total})

# ── Emp Stats ────────────────────────────────────────────────

@app.route("/api/emp-stats")
@require_admin
def emp_stats():
    period = request.args.get("period","month")
    if   period=="day":   cond="ts LIKE ?"; val=datetime.now().strftime("%Y-%m-%d")+"%"
    elif period=="week":  cond="ts>=?";     val=(datetime.now()-timedelta(days=7)).strftime("%Y-%m-%d")
    else:                 cond="ts LIKE ?"; val=datetime.now().strftime("%Y-%m")+"%"
    db   = get_db()
    rows = db.execute(f"""SELECT username,user_group,action,COUNT(*) as cnt FROM audit_logs
        WHERE {cond} AND action IN ('إضافة','تعديل','مغادرة','حذف')
        GROUP BY username,user_group,action""",(val,)).fetchall()
    db.close()
    users={}
    for r in rows:
        u=r["username"]
        if u not in users:
            users[u]={"username":u,"group":r["user_group"],"add":0,"edit":0,"depart":0,"delete":0}
        if r["action"]=="إضافة":   users[u]["add"]   +=r["cnt"]
        elif r["action"]=="تعديل": users[u]["edit"]  +=r["cnt"]
        elif r["action"]=="مغادرة":users[u]["depart"]+=r["cnt"]
        elif r["action"]=="حذف":   users[u]["delete"]+=r["cnt"]
    return jsonify({"ok":True,"stats":sorted(users.values(),
                    key=lambda x:x["add"]+x["edit"]+x["depart"],reverse=True)})

# ── Users ────────────────────────────────────────────────────

@app.route("/api/users")
@require_admin
def get_users():
    db    = get_db()
    users = db.execute("""SELECT u.id,u.username,u.role,u.group_id,u.active,g.name as group_name
        FROM users u LEFT JOIN groups g ON u.group_id=g.id ORDER BY u.role DESC,u.username""").fetchall()
    db.close()
    return jsonify({"ok":True,"users":[dict(u) for u in users]})

@app.route("/api/users/save", methods=["POST"])
@require_admin
def save_user():
    d        = request.json or {}
    uid      = d.get("id")
    username = d.get("username","").strip()
    password = d.get("password","")
    role     = d.get("role","worker")
    group_id = d.get("group_id","") or None
    active   = int(d.get("active",1))
    if role == "worker" and not group_id:
        return jsonify({"ok":False,"msg":"الموظف يجب أن يكون منسوباً لمجموعة"})
    db = get_db()
    try:
        if uid:
            if password:
                db.execute("UPDATE users SET role=?,group_id=?,active=?,password_hash=? WHERE id=?",
                           (role,group_id,active,generate_password_hash(password),uid))
            else:
                db.execute("UPDATE users SET role=?,group_id=?,active=? WHERE id=?",(role,group_id,active,uid))
        else:
            if not username or not password:
                db.close(); return jsonify({"ok":False,"msg":"اسم المستخدم وكلمة السر مطلوبان"})
            db.execute("INSERT INTO users (username,password_hash,role,group_id,active) VALUES (?,?,?,?,?)",
                       (username,generate_password_hash(password),role,group_id,active))
        db.commit()
    except sqlite3.IntegrityError:
        db.close(); return jsonify({"ok":False,"msg":"اسم المستخدم موجود مسبقاً"})
    db.close()
    return jsonify({"ok":True})

@app.route("/api/users/delete", methods=["POST"])
@require_admin
def delete_user():
    uid = (request.json or {}).get("id")
    db  = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?",(uid,)).fetchone()
    if not user: db.close(); return jsonify({"ok":False,"msg":"المستخدم غير موجود"})
    if user["username"] == "admin": db.close(); return jsonify({"ok":False,"msg":"لا يمكن حذف حساب admin"})
    if str(user["id"]) == str(session.get("user_id")):
        db.close(); return jsonify({"ok":False,"msg":"لا يمكن حذف حسابك الحالي"})
    db.execute("DELETE FROM users WHERE id=?",(uid,))
    db.commit(); db.close()
    return jsonify({"ok":True})

# ── Groups ───────────────────────────────────────────────────

@app.route("/api/groups")
@require_admin
def get_groups():
    db     = get_db()
    groups = db.execute("SELECT * FROM groups ORDER BY name").fetchall()
    db.close()
    return jsonify({"ok":True,"groups":[dict(g) for g in groups]})

@app.route("/api/groups/save", methods=["POST"])
@require_admin
def save_group():
    d      = request.json or {}
    gid    = d.get("id","").strip().lower()
    name   = d.get("name","").strip()
    prefix = d.get("prefix","").strip().upper()
    orig   = d.get("orig_id","")
    if not gid or not name:
        return jsonify({"ok":False,"msg":"المعرف والاسم مطلوبان"})
    if not prefix:
        return jsonify({"ok":False,"msg":"البادئة مطلوبة"})
    db = get_db()
    # check prefix uniqueness (exclude self)
    dup = db.execute("SELECT id FROM groups WHERE UPPER(prefix)=? AND id!=?",
                     (prefix, orig or gid)).fetchone()
    if dup:
        db.close()
        return jsonify({"ok":False,"msg":f"البادئة '{prefix}' مستخدمة من مجموعة أخرى"})
    try:
        if orig and orig != gid:
            db.execute("UPDATE pilgrims SET group_id=? WHERE group_id=?",(gid,orig))
            db.execute("UPDATE users    SET group_id=? WHERE group_id=?",(gid,orig))
            db.execute("DELETE FROM groups WHERE id=?",(orig,))
        db.execute("INSERT OR REPLACE INTO groups (id,name,prefix) VALUES (?,?,?)",(gid,name,prefix))
        db.commit()
    except Exception as e:
        db.close(); return jsonify({"ok":False,"msg":str(e)})
    db.close()
    log_action("تعديل مجموعة", new={"id":gid,"name":name,"prefix":prefix})
    return jsonify({"ok":True})

@app.route("/api/groups/delete", methods=["POST"])
@require_admin
def delete_group():
    d          = request.json or {}
    gid        = d.get("id","")
    emp_action = d.get("emp_action","disable")  # "disable" or "delete"
    db  = get_db()
    grp = db.execute("SELECT * FROM groups WHERE id=?",(gid,)).fetchone()
    if not grp: db.close(); return jsonify({"ok":False,"msg":"المجموعة غير موجودة"})
    if emp_action == "delete":
        db.execute("DELETE FROM users WHERE group_id=? AND role='worker'",(gid,))
    else:
        db.execute("UPDATE users SET active=0 WHERE group_id=?",(gid,))
    db.execute("DELETE FROM groups WHERE id=?",(gid,))
    db.commit(); db.close()
    log_action("حذف مجموعة", old={"id":gid,"name":grp["name"]})
    return jsonify({"ok":True})

# ── Backup ───────────────────────────────────────────────────

@app.route("/api/backup", methods=["POST"])
@require_admin
def backup():
    try:
        path = do_backup()
        return jsonify({"ok":True,"file":os.path.basename(path)})
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)})

@app.route("/api/backups")
@require_admin
def list_backups():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    files = sorted((f for f in os.listdir(BACKUP_DIR) if f.endswith(".db")), reverse=True)
    return jsonify({"ok":True,"files":files})

# ── Export ───────────────────────────────────────────────────

@app.route("/api/export")
@require_login
def export():
    db   = get_db()
    grps = db.execute("SELECT * FROM groups").fetchall()
    if session["role"] != "admin":
        grps = [g for g in grps if g["id"] == session["group_id"]]
    wb   = Workbook(); wb.remove(wb.active)
    NV="1A3A5C";BL="2563A8";WH="FFFFFF";GR="166534";GRB="DCFCE7"
    RD="991B1B";RDB="FEE2E2";OR="9A3412";ORB="FFEDD5";AL="F0F5FB"
    def F(c): return PatternFill("solid",fgColor=c)
    def B(c="CBD5E1"):
        s=Side(style='thin',color=c); return Border(left=s,right=s,top=s,bottom=s)
    def fnt(sz=10,b=False,c="1E293B"): return Font(name="Calibri",size=sz,bold=b,color=c)
    def aln(h='center',v='center'): return Alignment(horizontal=h,vertical=v)
    ws=wb.create_sheet("ملخص"); ws.sheet_view.rightToLeft=True
    for col,w in [('A',20),('B',12),('C',12),('D',12),('E',12),('F',14)]:
        ws.column_dimensions[col].width=w
    ws.merge_cells('A1:F1'); c=ws['A1']
    c.value=f"🕋 ملخص الجوازات — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font=Font(name="Calibri",size=14,bold=True,color=WH); c.fill=F(NV); c.alignment=aln()
    ws.row_dimensions[1].height=36
    for i,h in enumerate(["المجموعة","الإجمالي","مفعّل","غير مفعّل","غادر","نسبة المغادرة"],1):
        c=ws.cell(row=2,column=i,value=h)
        c.font=Font(name="Calibri",size=10,bold=True,color=WH); c.fill=F(BL); c.alignment=aln(); c.border=B(WH)
    ws.row_dimensions[2].height=24
    grand=[0,0,0,0]
    for ri,g in enumerate(grps):
        rows=db.execute("SELECT status,departed FROM pilgrims WHERE group_id=?",(g["id"],)).fetchall()
        t=len(rows); a=sum(1 for r in rows if r["status"]=="مفعل")
        iv=sum(1 for r in rows if r["status"]=="غير مفعل")
        dp=sum(1 for r in rows if r["departed"]=="نعم")
        grand[0]+=t;grand[1]+=a;grand[2]+=iv;grand[3]+=dp
        bg=AL if ri%2==0 else WH; rn=3+ri; ws.row_dimensions[rn].height=22
        ws.cell(rn,1,g["name"]).font=fnt(11,True,NV)
        ws.cell(rn,1).fill=F(bg); ws.cell(rn,1).alignment=aln('right'); ws.cell(rn,1).border=B()
        for ci,val in enumerate([t,a,iv,dp],2):
            c=ws.cell(rn,ci,val); c.fill=F(bg); c.alignment=aln(); c.border=B()
            clrs=[(NV,bg),(GR,GRB),(RD,RDB),(OR,ORB)]
            c.font=Font(name="Calibri",size=12,bold=True,color=clrs[ci-2][0]); c.fill=F(clrs[ci-2][1])
        pct=f"{int(dp/t*100)}%" if t>0 else "—"
        c=ws.cell(rn,6,pct); c.fill=F(bg); c.alignment=aln(); c.border=B(); c.font=fnt(11,True,NV)
    tr=3+len(grps); ws.row_dimensions[tr].height=28
    ws.cell(tr,1,"الإجمالي الكلي").font=Font(name="Calibri",size=11,bold=True,color=WH)
    ws.cell(tr,1).fill=F(NV); ws.cell(tr,1).alignment=aln('right'); ws.cell(tr,1).border=B(NV)
    for ci,val in enumerate(grand,2):
        c=ws.cell(tr,ci,val); c.font=Font(name="Calibri",size=13,bold=True,color=WH)
        c.fill=F(NV); c.alignment=aln(); c.border=B(NV)
    pct=f"{int(grand[3]/grand[0]*100)}%" if grand[0]>0 else "—"
    c=ws.cell(tr,6,pct); c.font=Font(name="Calibri",size=12,bold=True,color=WH)
    c.fill=F(NV); c.alignment=aln(); c.border=B(NV)
    for g in grps:
        ws2=wb.create_sheet(g["name"]); ws2.sheet_view.rightToLeft=True
        rows=db.execute("SELECT * FROM pilgrims WHERE group_id=? ORDER BY seq",(g["id"],)).fetchall()
        for cw,w in [(1,22),(2,8),(3,16),(4,12),(5,16),(6,30),(7,16)]:
            ws2.column_dimensions[chr(64+cw)].width=w
        ws2.merge_cells('A1:G1'); c=ws2['A1']; c.value=f"📋 {g['name']}"
        c.font=Font(name="Calibri",size=13,bold=True,color=WH); c.fill=F(NV); c.alignment=aln()
        ws2.row_dimensions[1].height=32
        for ci,h in enumerate(["رقم الجواز","م","حالة نسك","مغادر","وقت المغادرة","ملاحظات","آخر تعديل"],1):
            c=ws2.cell(2,ci,h); c.font=Font(name="Calibri",size=10,bold=True,color=WH)
            c.fill=F(BL); c.alignment=aln(); c.border=B(WH)
        ws2.row_dimensions[2].height=24
        for ri,r in enumerate(rows):
            rn=3+ri; bg=AL if ri%2==0 else WH; ws2.row_dimensions[rn].height=20
            for ci,val in enumerate([r["passport"],r["seq"],r["status"],r["departed"],
                                      r["departed_at"] or "",r["notes"] or "",r["updated_at"] or ""],1):
                c=ws2.cell(rn,ci,val); c.fill=F(bg); c.alignment=aln(); c.border=B()
                if ci==3:
                    if val=="مفعل": c.fill=F(GRB); c.font=Font(name="Calibri",bold=True,color=GR)
                    elif val=="غير مفعل": c.fill=F(RDB); c.font=Font(name="Calibri",bold=True,color=RD)
                elif ci==4 and val=="نعم":
                    c.fill=F(ORB); c.font=Font(name="Calibri",bold=True,color=OR)
    db.close()
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"جوازات_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── HTML ─────────────────────────────────────────────────────

@app.route("/")
def index():
    return Response(HTML_PAGE, mimetype="text/html; charset=utf-8")

HTML_PAGE = r"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>نظام الجوازات</title>
<style>
:root{
  --bg:#f5f5f7;--surface:#fff;--surface2:#f5f5f7;
  --tx:#1d1d1f;--tx2:#6e6e73;--tx3:#aeaeb2;
  --accent:#0071e3;--accent-d:#0064cc;
  --ok:#28cd41;--ok-bg:#edfbf0;--ok-tx:#1a7f2e;
  --err:#ff3b30;--err-bg:#fff0ef;--err-tx:#c0392b;
  --warn:#ff9f0a;--warn-bg:#fff8ec;--warn-tx:#974900;
  --border:#d2d2d7;--border-l:#e8e8ed;
  --shadow:0 1px 3px rgba(0,0,0,.08),0 4px 16px rgba(0,0,0,.05);
  --shadow-sm:0 1px 2px rgba(0,0,0,.06);
  --r:12px;--r-sm:8px;--r-lg:18px;
}
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Helvetica,Arial,sans-serif;
  background:var(--bg);color:var(--tx);font-size:15px;min-height:100vh;line-height:1.5}

/* ── Screens ── */
.scr{display:none;flex-direction:column;min-height:100vh}
.scr.on{display:flex!important}

/* ── Navbar ── */
.nav{height:56px;background:rgba(245,245,247,.88);backdrop-filter:blur(20px);
  -webkit-backdrop-filter:blur(20px);border-bottom:1px solid var(--border-l);
  display:flex;align-items:center;justify-content:space-between;
  padding:0 20px;position:sticky;top:0;z-index:100}
.nav-title{font-size:16px;font-weight:600;color:var(--tx)}
.nav-sub{font-size:12px;color:var(--tx2);margin-top:1px}
.nav-actions{display:flex;gap:8px;align-items:center}
.nav-btn{background:none;border:1px solid var(--border);border-radius:var(--r-sm);
  color:var(--tx);padding:6px 14px;font-size:13px;font-weight:500;cursor:pointer;
  font-family:inherit;transition:background .12s}
.nav-btn:hover{background:var(--border-l)}
.nav-btn.primary{background:var(--accent);border-color:var(--accent);color:#fff}
.nav-btn.primary:hover{background:var(--accent-d)}

/* ── Page body ── */
.pg{flex:1;padding:20px;max-width:580px;margin:0 auto;width:100%}
.pg-wide{flex:1;padding:20px;max-width:960px;margin:0 auto;width:100%}

/* ── Cards ── */
.card{background:var(--surface);border-radius:var(--r);border:1px solid var(--border-l);
  box-shadow:var(--shadow-sm);padding:20px;margin-bottom:14px}
.card-title{font-size:11px;font-weight:700;color:var(--tx2);text-transform:uppercase;
  letter-spacing:.5px;margin-bottom:16px}

/* ── Stat grid ── */
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:18px}
.stat{background:var(--surface);border-radius:var(--r);border:1px solid var(--border-l);
  padding:16px 12px;text-align:center}
.stat-n{font-size:28px;font-weight:700;line-height:1;color:var(--tx)}
.stat-l{font-size:11px;color:var(--tx2);margin-top:4px}
.c-ok{color:var(--ok-tx)}.c-err{color:var(--err-tx)}.c-warn{color:var(--warn-tx)}.c-accent{color:var(--accent)}

/* ── Form ── */
.field{margin-bottom:14px}
.field label{display:block;font-size:13px;font-weight:500;color:var(--tx2);margin-bottom:6px}
.inp{width:100%;padding:11px 14px;border:1.5px solid var(--border);border-radius:var(--r-sm);
  font-size:15px;color:var(--tx);background:var(--surface);outline:none;
  font-family:inherit;transition:border .12s,box-shadow .12s}
.inp:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(0,113,227,.1)}
.inp::placeholder{color:var(--tx3)}
.inp[type=password]{direction:ltr;letter-spacing:3px}
select.inp option{color:var(--tx)}
textarea.inp{resize:vertical;min-height:72px}
.msg-err{font-size:13px;color:var(--err-tx);margin-top:5px;display:none}
.msg-ok{font-size:13px;color:var(--ok-tx);margin-top:5px;display:none}

/* ── Buttons ── */
.btn{display:block;width:100%;padding:13px;border-radius:var(--r-sm);
  border:1.5px solid var(--border);background:var(--surface);
  color:var(--tx);font-size:15px;font-weight:600;cursor:pointer;
  text-align:center;transition:all .12s;font-family:inherit}
.btn:active{transform:scale(.98)}.btn:disabled{opacity:.45;cursor:not-allowed;transform:none}
.btn-p{background:var(--accent);border-color:var(--accent);color:#fff}
.btn-p:hover:not(:disabled){background:var(--accent-d)}
.btn-ok{background:var(--ok);border-color:var(--ok);color:#fff}
.btn-ok:hover:not(:disabled){background:#22b536}
.btn-err{background:var(--err);border-color:var(--err);color:#fff}
.btn-err:hover:not(:disabled){background:#e0352b}
.btn-warn{background:var(--warn);border-color:var(--warn);color:#fff}
.btn-g2{display:grid;grid-template-columns:1fr 1fr;gap:10px}

/* ── Badges ── */
.badge{display:inline-flex;align-items:center;padding:2px 10px;border-radius:99px;
  font-size:12px;font-weight:600}
.b-ok{background:var(--ok-bg);color:var(--ok-tx)}
.b-err{background:var(--err-bg);color:var(--err-tx)}
.b-warn{background:var(--warn-bg);color:var(--warn-tx)}
.b-blue{background:#e8f1fd;color:#0055a5}
.b-gray{background:#f0f0f5;color:var(--tx2)}
.b-dipl{background:rgba(109,15,44,.85);color:#fff}
.list-row.dipl{background:rgba(109,15,44,.08)}

/* ── Toggle ── */
.tog{position:relative;width:48px;height:26px;flex-shrink:0}
.tog input{opacity:0;width:0;height:0;position:absolute}
.tog-sl{position:absolute;inset:0;background:var(--border);border-radius:99px;
  cursor:pointer;transition:background .2s}
.tog-sl:before{content:"";position:absolute;width:20px;height:20px;background:#fff;
  border-radius:50%;top:3px;right:3px;transition:right .2s;
  box-shadow:0 1px 3px rgba(0,0,0,.2)}
.tog input:checked+.tog-sl{background:var(--ok)}
.tog input:checked+.tog-sl:before{right:25px}

/* ── Row items ── */
.row{display:flex;align-items:center;justify-content:space-between;
  padding:11px 0;border-bottom:1px solid var(--border-l)}
.row:last-child{border-bottom:none}
.row-l{font-size:13px;color:var(--tx2)}.row-v{font-size:14px;font-weight:500}
.trow{display:flex;align-items:center;justify-content:space-between;
  padding:12px 0;border-bottom:1px solid var(--border-l)}
.trow:last-child{border-bottom:none}

/* ── Result box ── */
.result-box{background:var(--surface2);border-radius:var(--r);
  padding:16px;margin-top:14px;border:1px solid var(--border-l)}

/* ── Group items (in admin list) ── */
.g-row{display:flex;align-items:center;justify-content:space-between;
  padding:14px 0;border-bottom:1px solid var(--border-l);flex-wrap:wrap;gap:8px}
.g-row:last-child{border-bottom:none}
.g-name{font-size:15px;font-weight:600}
.g-sub{font-size:12px;color:var(--tx2);margin-top:2px}

/* ── List rows ── */
.list-row{display:flex;align-items:center;justify-content:space-between;
  padding:12px 0;border-bottom:1px solid var(--border-l);cursor:pointer;gap:8px}
.list-row:last-child{border-bottom:none}
.list-row:hover{margin:0 -4px;padding-left:4px;padding-right:4px;
  background:var(--surface2);border-radius:var(--r-sm)}
.lp{font-size:14px;font-weight:700;direction:ltr}.ln{font-size:12px;color:var(--tx2);margin-top:1px}
.badge-row{display:flex;gap:4px;flex-wrap:wrap;justify-content:flex-end}

/* ── Filter bar ── */
.filters{display:flex;gap:7px;overflow-x:auto;padding-bottom:4px;
  margin-bottom:14px;scrollbar-width:none}
.fb{flex-shrink:0;padding:5px 14px;border-radius:99px;border:1.5px solid var(--border);
  background:var(--surface);font-size:13px;cursor:pointer;color:var(--tx2);
  font-family:inherit;transition:all .12s}
.fb.on{background:var(--accent);color:#fff;border-color:var(--accent)}

/* ── Table ── */
.tbl-wrap{overflow-x:auto;border-radius:var(--r-sm)}
.tbl{width:100%;border-collapse:collapse;font-size:13px}
.tbl th{background:var(--surface2);color:var(--tx2);padding:10px 12px;
  text-align:right;font-weight:600;border-bottom:1px solid var(--border-l);white-space:nowrap}
.tbl td{padding:9px 12px;border-bottom:1px solid var(--border-l);color:var(--tx)}
.tbl tr:last-child td{border-bottom:none}
.tbl tr:hover td{background:var(--surface2)}

/* ── Pagination ── */
.pag{display:flex;align-items:center;justify-content:center;gap:8px;margin-top:14px}
.pag button{background:var(--surface);border:1px solid var(--border);
  color:var(--tx);border-radius:var(--r-sm);padding:6px 14px;cursor:pointer;
  font-family:inherit;font-size:13px}
.pag button:disabled{opacity:.35;cursor:not-allowed}
.pag span{font-size:13px;color:var(--tx2)}

/* ── Tabs ── */
.tabs{display:flex;gap:4px;overflow-x:auto;padding-bottom:2px;
  margin-bottom:20px;scrollbar-width:none}
.tab{flex-shrink:0;padding:8px 16px;border-radius:99px;border:1.5px solid var(--border-l);
  background:var(--surface);font-size:13px;cursor:pointer;color:var(--tx2);
  font-family:inherit;font-weight:500;transition:all .12s}
.tab.on{background:var(--accent);color:#fff;border-color:var(--accent)}

/* ── Login ── */
.login-wrap{min-height:100vh;display:flex;align-items:center;
  justify-content:center;background:var(--bg);padding:20px}
.login-card{width:100%;max-width:400px;background:var(--surface);
  border-radius:var(--r-lg);padding:36px 28px;box-shadow:var(--shadow);
  border:1px solid var(--border-l)}
.login-icon{width:68px;height:68px;background:var(--bg);border-radius:16px;
  display:flex;align-items:center;justify-content:center;margin:0 auto 18px;
  font-size:32px;border:1px solid var(--border-l)}
.login-title{text-align:center;font-size:22px;font-weight:700;margin-bottom:6px}
.login-sub{text-align:center;font-size:14px;color:var(--tx2);margin-bottom:24px}
.groups-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:16px}
.g-card{background:var(--bg);border:1.5px solid var(--border-l);border-radius:var(--r);
  padding:14px 10px;text-align:center;cursor:pointer;transition:all .15s}
.g-card:hover{border-color:var(--accent);background:#eef6ff}
.g-card-name{font-size:14px;font-weight:600}
.g-card-ico{font-size:20px;margin-bottom:6px}
.divider{display:flex;align-items:center;gap:10px;margin:14px 0}
.divider::before,.divider::after{content:"";flex:1;height:1px;background:var(--border-l)}
.divider span{font-size:12px;color:var(--tx3)}
.back-link{background:none;border:none;color:var(--accent);font-size:14px;
  cursor:pointer;font-family:inherit;padding:0;margin-bottom:18px;display:inline-flex;
  align-items:center;gap:4px}
.selected-badge{display:inline-block;background:#eef6ff;color:var(--accent);
  border:1px solid #b8d9f8;border-radius:99px;padding:4px 14px;
  font-size:13px;font-weight:600;margin-bottom:20px}

/* ── Modal ── */
.modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.35);
  backdrop-filter:blur(4px);-webkit-backdrop-filter:blur(4px);
  z-index:500;align-items:center;justify-content:center;padding:20px}
.modal-bg.open{display:flex}
.modal{background:var(--surface);border-radius:var(--r-lg);padding:28px;
  width:100%;max-width:420px;box-shadow:0 20px 60px rgba(0,0,0,.15);
  border:1px solid var(--border-l)}
.modal h3{font-size:18px;font-weight:700;margin-bottom:18px}
.modal-actions{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:20px}

/* ── Toast ── */
.toast{position:fixed;bottom:28px;left:50%;transform:translateX(-50%) translateY(100px);
  padding:12px 24px;border-radius:99px;font-size:14px;font-weight:600;
  transition:transform .28s cubic-bezier(.34,1.56,.64,1);z-index:9999;
  pointer-events:none;color:#fff;white-space:nowrap;
  box-shadow:0 4px 20px rgba(0,0,0,.2)}
.toast.show{transform:translateX(-50%) translateY(0)}

/* ── Spinner ── */
.sp{display:inline-block;width:14px;height:14px;border:2px solid rgba(255,255,255,.4);
  border-top-color:#fff;border-radius:50%;animation:sp .7s linear infinite;
  vertical-align:middle;margin-left:6px}
@keyframes sp{to{transform:rotate(360deg)}}

/* ── Departure card ── */
.dep-card{border:1.5px solid var(--warn-bg);background:var(--warn-bg)}
.dep-done{border-color:var(--ok-bg);background:var(--ok-bg)}

/* ── Responsive ── */
@media(max-width:520px){
  .stats{grid-template-columns:1fr 1fr}
  .pg,.pg-wide{padding:14px}
  .login-card{padding:28px 20px}
  .groups-grid{grid-template-columns:1fr 1fr}
}
</style>
</head>
<body>

<!-- ══ LOGIN ══════════════════════════════════════════════════ -->
<div class="scr on" id="s-login">
<div class="login-wrap">

  <!-- Step 1 -->
  <div class="login-card" id="step1">
    <div class="login-icon">🕋</div>
    <div class="login-title">نظام تنظيم الجوازات</div>
    <div class="login-sub">اختر مجموعتك للمتابعة</div>
    <div class="groups-grid" id="grp-grid">
      <div style="grid-column:1/-1;text-align:center;padding:16px;color:var(--tx3)">جاري التحميل...</div>
    </div>
    <div class="divider"><span>أو</span></div>
    <button class="btn" style="font-size:14px" onclick="pickGroup('admin','مشرف عام')">🔐 دخول المشرف</button>
  </div>

  <!-- Step 2 -->
  <div class="login-card" id="step2" style="display:none">
    <button class="back-link" onclick="backToStep1()">← رجوع</button>
    <div><span class="selected-badge" id="sel-badge"></span></div>
    <div class="field" style="margin-top:18px"><label>اسم المستخدم</label>
      <input class="inp" type="text" id="l-user" placeholder="username" autocomplete="username" style="direction:ltr">
    </div>
    <div class="field"><label>كلمة السر</label>
      <input class="inp" type="password" id="l-pw" placeholder="••••••"
             onkeydown="if(event.key==='Enter')doLogin()">
      <div class="msg-err" id="l-err"></div>
    </div>
    <button class="btn btn-p" id="l-btn" onclick="doLogin()">دخول</button>
  </div>

</div>
</div>

<!-- ══ ADMIN ══════════════════════════════════════════════════ -->
<div class="scr" id="s-admin">
  <nav class="nav">
    <div><div class="nav-title">لوحة التحكم</div><div class="nav-sub">مشرف عام</div></div>
    <div class="nav-actions">
      <button class="nav-btn primary" onclick="showDepart()">✈️ مغادرة</button>
      <button class="nav-btn" onclick="exportExcel()">📥 تصدير</button>
      <button class="nav-btn" onclick="doLogout()">خروج</button>
    </div>
  </nav>
  <div class="pg-wide">
    <div class="stats">
      <div class="stat"><div class="stat-n c-accent" id="at">—</div><div class="stat-l">إجمالي</div></div>
      <div class="stat"><div class="stat-n c-ok"     id="aa">—</div><div class="stat-l">مفعّل</div></div>
      <div class="stat"><div class="stat-n c-warn"   id="ad">—</div><div class="stat-l">غادروا</div></div>
      <div class="stat"><div class="stat-n c-err"    id="ai">—</div><div class="stat-l">غير مفعّل</div></div>
    </div>
    <div class="tabs">
      <button class="tab on" onclick="aTab('ov',this)">نظرة عامة</button>
      <button class="tab" onclick="aTab('logs',this)">السجلات</button>
      <button class="tab" onclick="aTab('perf',this)">الأداء</button>
      <button class="tab" onclick="aTab('emp',this)">الموظفين</button>
      <button class="tab" onclick="aTab('grp',this)">المجموعات</button>
      <button class="tab" onclick="aTab('bak',this)">النسخ</button>
    </div>

    <!-- Overview -->
    <div id="t-ov">
      <div class="card"><div class="card-title">المجموعات</div>
        <div id="adm-grps"><div style="text-align:center;padding:20px;color:var(--tx3)">جاري التحميل...</div></div>
      </div>
      <div class="card"><div class="card-title">بحث شامل</div>
        <div class="field" style="margin-bottom:10px">
          <input class="inp" type="text" id="gs" placeholder="رقم الجواز..." style="direction:ltr;letter-spacing:1px"
                 onkeydown="if(event.key==='Enter')doGSearch()">
        </div>
        <button class="btn btn-p" id="gs-btn" onclick="doGSearch()">بحث في جميع المجموعات</button>
        <div class="result-box" id="gr" style="display:none">
          <div class="row"><span class="row-l">الرقم التسلسلي</span>
            <span class="row-v" style="color:var(--accent);font-weight:700;direction:ltr" id="gr-seq"></span></div>
          <div class="row"><span class="row-l">المجموعة</span><span class="row-v" id="gr-g"></span></div>
          <div class="row"><span class="row-l">حالة النسك</span><span id="gr-st"></span></div>
          <div class="row"><span class="row-l">مغادر</span><span id="gr-dp"></span></div>
          <div class="row"><span class="row-l">وقت المغادرة</span><span class="row-v" id="gr-dat"></span></div>
          <div class="row" style="border:none"><span class="row-l">ملاحظات</span>
            <span class="row-v" style="font-size:12px;max-width:55%;text-align:left" id="gr-nt"></span></div>
          <div class="row" style="border:none;margin-top:10px">
            <span class="row-l">دبلوماسي</span>
            <span id="gr-dipl-badge"></span>
            <button class="nav-btn" id="gr-dipl-btn" onclick="toggleDiplomatic()" style="margin-right:8px;font-size:12px;padding:4px 10px"></button>
          </div>
        </div>
      </div>
      <button class="btn" onclick="loadAdmin()">↺ تحديث</button>
    </div>

    <!-- Logs -->
    <div id="t-logs" style="display:none">
      <div class="card"><div class="card-title">فلتر</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px">
          <div class="field" style="margin:0"><label>المستخدم</label>
            <input class="inp" type="text" id="lf-u" placeholder="الكل"></div>
          <div class="field" style="margin:0"><label>العملية</label>
            <select class="inp" id="lf-a">
              <option value="">الكل</option>
              <option>دخول</option><option>خروج</option>
              <option>إضافة</option><option>تعديل</option>
              <option>مغادرة</option><option>حذف</option>
            </select></div>
        </div>
        <div class="field" style="margin-bottom:10px"><label>رقم الجواز</label>
          <input class="inp" type="text" id="lf-p" placeholder="بحث..." style="direction:ltr"></div>
        <button class="btn btn-p" onclick="loadLogs(1)">بحث</button>
      </div>
      <div class="card">
        <div class="tbl-wrap">
          <table class="tbl">
            <thead><tr><th>الوقت</th><th>المستخدم</th><th>العملية</th><th>الجواز</th></tr></thead>
            <tbody id="logs-body">
              <tr><td colspan="4" style="text-align:center;padding:20px;color:var(--tx3)">اضغط بحث</td></tr>
            </tbody>
          </table>
        </div>
        <div class="pag" id="logs-pag"></div>
      </div>
    </div>

    <!-- Performance -->
    <div id="t-perf" style="display:none">
      <div class="card"><div class="card-title">فترة الإحصاء</div>
        <div class="filters">
          <button class="fb on" onclick="loadPerf('day',this)">اليوم</button>
          <button class="fb" onclick="loadPerf('week',this)">الأسبوع</button>
          <button class="fb" onclick="loadPerf('month',this)">الشهر</button>
        </div>
        <div class="tbl-wrap">
          <table class="tbl">
            <thead><tr><th>#</th><th>الموظف</th><th>المجموعة</th><th>إضافات</th><th>تعديلات</th><th>مغادرات</th><th>حذف</th><th>المجموع</th></tr></thead>
            <tbody id="perf-body">
              <tr><td colspan="8" style="text-align:center;padding:20px;color:var(--tx3)">جاري التحميل...</td></tr>
            </tbody>
          </table>
        </div>
      </div>
    </div>

    <!-- Employees -->
    <div id="t-emp" style="display:none">
      <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
          <span class="card-title" style="margin:0">الموظفون</span>
          <button class="nav-btn primary" onclick="openUserModal(null)">+ إضافة</button>
        </div>
        <div id="emp-list"><div style="text-align:center;padding:20px;color:var(--tx3)">جاري التحميل...</div></div>
      </div>
    </div>

    <!-- Groups Mgmt -->
    <div id="t-grp" style="display:none">
      <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
          <span class="card-title" style="margin:0">المجموعات</span>
          <button class="nav-btn primary" onclick="openGroupModal(null)">+ إضافة</button>
        </div>
        <div id="grp-list"><div style="text-align:center;padding:20px;color:var(--tx3)">جاري التحميل...</div></div>
      </div>
    </div>

    <!-- Backup -->
    <div id="t-bak" style="display:none">
      <div class="card"><div class="card-title">النسخ الاحتياطي</div>
        <button class="btn btn-p" id="bak-btn" onclick="doBak()" style="margin-bottom:16px">💾 إنشاء نسخة الآن</button>
        <div id="bak-list"><div style="text-align:center;padding:16px;color:var(--tx3)">جاري التحميل...</div></div>
      </div>
    </div>
  </div>
</div>

<!-- ══ WORKER ══════════════════════════════════════════════════ -->
<div class="scr" id="s-worker">
  <nav class="nav">
    <div><div class="nav-title" id="w-name">المجموعة</div><div class="nav-sub">موظف</div></div>
    <div class="nav-actions">
      <button class="nav-btn primary" onclick="showDepart()">✈️</button>
      <button class="nav-btn" onclick="showList()">📋</button>
      <button class="nav-btn" onclick="show('s-add')">➕</button>
      <button class="nav-btn" onclick="doLogout()">خروج</button>
    </div>
  </nav>
  <div class="pg">
    <div class="stats">
      <div class="stat"><div class="stat-n c-accent" id="wt">—</div><div class="stat-l">إجمالي</div></div>
      <div class="stat"><div class="stat-n c-ok"     id="wa">—</div><div class="stat-l">مفعّل</div></div>
      <div class="stat"><div class="stat-n c-warn"   id="wd">—</div><div class="stat-l">غادر</div></div>
      <div class="stat"><div class="stat-n c-err"    id="wi">—</div><div class="stat-l">غير مفعّل</div></div>
    </div>
    <div class="card"><div class="card-title">بحث عن حاج</div>
      <div class="field" style="margin-bottom:10px">
        <input class="inp" type="text" id="ws" placeholder="رقم الجواز..."
               style="direction:ltr;letter-spacing:1px"
               onkeydown="if(event.key==='Enter')doWSearch()">
      </div>
      <button class="btn btn-p" id="ws-btn" onclick="doWSearch()">بحث</button>
      <div class="result-box" id="wr" style="display:none">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">
          <div>
            <div style="font-size:18px;font-weight:700;direction:ltr" id="wr-pp"></div>
            <div id="wr-dipl" style="margin-top:5px;display:none"></div>
          </div>
          <span class="badge" id="wr-tag"></span>
        </div>
        <div class="row"><span class="row-l">التسلسل</span><span class="row-v" id="wr-seq"></span></div>
        <div class="row"><span class="row-l">ملاحظات</span>
          <span class="row-v" style="font-size:12px;max-width:55%;text-align:left" id="wr-nt"></span></div>
        <div style="margin:14px 0 8px">
          <div class="trow"><span style="font-size:15px">بطاقة نسك مفعّلة</span>
            <label class="tog"><input type="checkbox" id="t-st"><span class="tog-sl"></span></label></div>
        </div>
        <div class="field"><label>ملاحظات</label>
          <textarea class="inp" id="w-notes" rows="2" placeholder="أي ملاحظة..."></textarea></div>
        <div class="btn-g2">
          <button class="btn btn-p" id="sv-btn" onclick="doSave()">حفظ</button>
          <button class="btn btn-err" onclick="doDelete()">حذف</button>
        </div>
      </div>
    </div>
    <div class="btn-g2">
      <button class="btn btn-ok" onclick="show('s-add')">➕ إضافة حاج</button>
      <button class="btn" onclick="showList()">📋 القائمة</button>
    </div>
  </div>
</div>

<!-- ══ ADD ════════════════════════════════════════════════════ -->
<div class="scr" id="s-add">
  <nav class="nav">
    <div><div class="nav-title">إضافة حاج جديد</div></div>
    <button class="nav-btn" onclick="goBack()">رجوع</button>
  </nav>
  <div class="pg"><div class="card">
    <div class="field"><label>رقم الجواز</label>
      <input class="inp" type="text" id="a-pp" placeholder="AA0000000"
             style="direction:ltr;letter-spacing:1px;font-size:18px"
             oninput="this.value=this.value.toUpperCase().replace(/[^A-Z0-9]/g,'')"
             onkeydown="if(event.key==='Enter')doAdd()">
      <div class="msg-err" id="a-err"></div><div class="msg-ok" id="a-ok"></div>
    </div>
    <div class="field"><label>حالة بطاقة نسك</label>
      <select class="inp" id="a-st">
        <option value="">— اختر —</option>
        <option value="مفعل">مفعّل ✓</option>
        <option value="غير مفعل">غير مفعّل ✗</option>
      </select>
    </div>
    <div class="field"><label>ملاحظات (اختياري)</label>
      <textarea class="inp" id="a-nt" rows="2" placeholder="أي ملاحظة..."></textarea>
    </div>
    <button class="btn btn-p" id="a-btn" onclick="doAdd()">➕ إضافة</button>
  </div></div>
</div>

<!-- ══ LIST ═══════════════════════════════════════════════════ -->
<div class="scr" id="s-list">
  <nav class="nav">
    <div><div class="nav-title">قائمة الحجاج</div><div class="nav-sub" id="list-cnt"></div></div>
    <button class="nav-btn" onclick="goBack()">رجوع</button>
  </nav>
  <div class="pg">
    <div class="filters">
      <button class="fb on" onclick="setF('all',this)">الكل</button>
      <button class="fb" onclick="setF('active',this)">مفعّل</button>
      <button class="fb" onclick="setF('inactive',this)">غير مفعّل</button>
      <button class="fb" onclick="setF('departed',this)">غادر</button>
      <button class="fb" onclick="setF('pending',this)">لم يغادر</button>
    </div>
    <div class="field" style="margin-bottom:10px">
      <input class="inp" type="text" id="list-q" placeholder="ابحث في القائمة..." oninput="fLocal()">
    </div>
    <div class="card" style="padding:0 18px">
      <div id="list-rows"><div style="text-align:center;padding:24px;color:var(--tx3)">جاري التحميل...</div></div>
    </div>
    <div class="pag" id="list-pag"></div>
  </div>
</div>

<!-- ══ DEPART ════════════════════════════════════════════════ -->
<div class="scr" id="s-depart">
  <nav class="nav">
    <div><div class="nav-title">✈️ صفحة المغادرة</div></div>
    <button class="nav-btn" onclick="goBack()">رجوع</button>
  </nav>
  <div class="pg">
    <div class="card"><div class="card-title">بحث برقم الجواز</div>
      <div class="field" style="margin-bottom:10px">
        <input class="inp" type="text" id="dp-q" placeholder="رقم الجواز..."
               style="direction:ltr;letter-spacing:1px;font-size:17px"
               onkeydown="if(event.key==='Enter')doDSearch()">
      </div>
      <button class="btn btn-warn" id="dp-btn" onclick="doDSearch()">🔍 بحث</button>
    </div>
    <div class="card dep-card" id="dp-res" style="display:none">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
        <div>
          <div style="font-size:22px;font-weight:700;direction:ltr" id="dp-pp"></div>
          <div style="font-size:12px;color:var(--tx2);margin-top:3px" id="dp-grp"></div>
          <div id="dp-dipl" style="margin-top:5px;display:none"></div>
        </div>
        <span class="badge" id="dp-tag"></span>
      </div>
      <div class="row"><span class="row-l">حالة النسك</span><span id="dp-st"></span></div>
      <div class="row" id="dp-time-r" style="display:none">
        <span class="row-l">وقت المغادرة</span><span class="row-v" id="dp-time"></span></div>
      <div class="row" style="border:none"><span class="row-l">ملاحظات</span>
        <span class="row-v" style="font-size:12px" id="dp-nt"></span></div>
      <div id="dp-action" style="margin-top:16px"></div>
    </div>
  </div>
</div>

<!-- ══ MODALS ════════════════════════════════════════════════ -->

<!-- User modal -->
<div class="modal-bg" id="m-user">
  <div class="modal">
    <h3 id="um-title">إضافة موظف</h3>
    <input type="hidden" id="um-id">
    <div class="field"><label>اسم المستخدم</label>
      <input class="inp" type="text" id="um-user" placeholder="username" style="direction:ltr">
    </div>
    <div class="field"><label>كلمة السر <span style="color:var(--tx3);font-size:11px">(فارغ = لا تغيير)</span></label>
      <input class="inp" type="password" id="um-pw" placeholder="••••••">
    </div>
    <div class="field"><label>الدور</label>
      <select class="inp" id="um-role" onchange="toggleGroupField()">
        <option value="worker">موظف</option>
        <option value="admin">مشرف</option>
      </select>
    </div>
    <div class="field" id="um-grp-f"><label>المجموعة</label>
      <select class="inp" id="um-grp"><option value="">— اختر —</option></select>
    </div>
    <div class="field"><label>الحالة</label>
      <select class="inp" id="um-active">
        <option value="1">مفعّل</option>
        <option value="0">معطّل</option>
      </select>
    </div>
    <div class="msg-err" id="um-err"></div>
    <div class="modal-actions">
      <button class="btn btn-p" onclick="saveUser()">حفظ</button>
      <button class="btn" onclick="close('m-user')">إلغاء</button>
    </div>
  </div>
</div>

<!-- Group modal -->
<div class="modal-bg" id="m-grp">
  <div class="modal">
    <h3 id="gm-title">إضافة مجموعة</h3>
    <input type="hidden" id="gm-orig">
    <div class="field"><label>معرّف المجموعة (إنجليزي)</label>
      <input class="inp" type="text" id="gm-id" placeholder="group_id" style="direction:ltr">
    </div>
    <div class="field"><label>اسم المجموعة</label>
      <input class="inp" type="text" id="gm-name" placeholder="الاسم بالعربي">
    </div>
    <div class="field"><label>البادئة <span style="color:var(--tx3);font-size:11px">(حرف أو حرفان — مثال: A)</span></label>
      <input class="inp" type="text" id="gm-prefix" placeholder="A"
             style="direction:ltr;letter-spacing:3px;font-weight:700;font-size:18px;max-width:80px"
             maxlength="3" oninput="this.value=this.value.toUpperCase()">
    </div>
    <div style="background:var(--surface2);border-radius:var(--r-sm);padding:12px;font-size:13px;color:var(--tx2);margin-bottom:14px">
      السجلات ستُرقَّم: <strong id="gm-preview" style="color:var(--accent);direction:ltr;display:inline-block">A-1، A-2...</strong>
    </div>
    <div class="msg-err" id="gm-err"></div>
    <div class="modal-actions">
      <button class="btn btn-p" onclick="saveGroup()">حفظ</button>
      <button class="btn" onclick="close('m-grp')">إلغاء</button>
    </div>
  </div>
</div>

<!-- Delete pilgrim confirm -->
<div class="modal-bg" id="m-del">
  <div class="modal">
    <h3>⚠️ تأكيد الحذف</h3>
    <p style="color:var(--tx2);margin-bottom:20px">
      هل أنت متأكد من حذف جواز <strong id="del-pp" style="color:var(--err-tx);direction:ltr;display:inline-block"></strong>؟<br>
      <small style="color:var(--tx3)">لا يمكن التراجع.</small>
    </p>
    <div class="modal-actions">
      <button class="btn btn-err" onclick="confirmDel()">نعم، احذف</button>
      <button class="btn" onclick="close('m-del')">إلغاء</button>
    </div>
  </div>
</div>

<!-- Delete group confirm -->
<div class="modal-bg" id="m-delgrp">
  <div class="modal">
    <h3>🗑️ حذف مجموعة</h3>
    <p style="color:var(--tx2);margin-bottom:16px">
      ستُحذف مجموعة <strong id="delgrp-name" style="color:var(--err-tx)"></strong> نهائياً.<br>
      <small style="color:var(--tx3)">الجوازات المرتبطة تبقى في قاعدة البيانات.</small>
    </p>
    <div class="field"><label>الموظفون المرتبطون:</label>
      <select class="inp" id="delgrp-emp">
        <option value="disable">تعطيل حسابات الموظفين</option>
        <option value="delete">حذف حسابات الموظفين</option>
      </select>
    </div>
    <input type="hidden" id="delgrp-id">
    <div class="modal-actions">
      <button class="btn btn-err" onclick="confirmDelGrp()">نعم، احذف</button>
      <button class="btn" onclick="close('m-delgrp')">إلغاء</button>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
var S=null, CR=null, listData=[], curF='all', listPage=1, logsPage=1, listGid='';

// ── Core ────────────────────────────────────────────────────
function api(url,opts){
  return fetch(url,opts).then(r=>{
    if(r.status===401){doLogout();return{ok:false,msg:'انتهت الجلسة'};}
    return r.json();
  }).catch(e=>({ok:false,msg:e.message}));
}
function show(id){
  document.querySelectorAll('.scr').forEach(s=>s.classList.remove('on'));
  document.getElementById(id).classList.add('on');
  window.scrollTo(0,0);
}
function goBack(){show(S&&S.role==='admin'?'s-admin':'s-worker');}
function close(id){document.getElementById(id).classList.remove('open');}
function toast(m,e){
  var t=document.getElementById('toast');
  t.textContent=m; t.style.background=e?'#c0392b':'#1a7f2e';
  t.classList.add('show'); setTimeout(()=>t.classList.remove('show'),3000);
}
function sb(id,t,l){
  var b=document.getElementById(id); if(!b)return;
  b.innerHTML=l?t+'<span class="sp"></span>':t; b.disabled=l;
}
function badge(txt,cls){return `<span class="badge ${cls}">${txt}</span>`;}

// ── Login ───────────────────────────────────────────────────
var selGid='', selGname='';

async function loadLoginGroups(){
  var r=await fetch('/api/public/groups').then(x=>x.json()).catch(()=>({ok:false}));
  var grid=document.getElementById('grp-grid');
  if(!r.ok||!r.groups.length){
    grid.innerHTML='<div style="grid-column:1/-1;text-align:center;color:var(--tx3);padding:12px">لا توجد مجموعات</div>';
    return;
  }
  grid.innerHTML=r.groups.map(g=>`
    <div class="g-card" onclick="pickGroup('${g.id}','${g.name}')">
      <div class="g-card-ico">👥</div>
      <div class="g-card-name">${g.name}</div>
    </div>`).join('');
}

function pickGroup(gid,name){
  selGid=gid; selGname=name;
  document.getElementById('sel-badge').textContent=name;
  document.getElementById('l-user').value='';
  document.getElementById('l-pw').value='';
  document.getElementById('l-err').style.display='none';
  document.getElementById('step1').style.display='none';
  document.getElementById('step2').style.display='block';
  document.getElementById('l-user').focus();
}
function backToStep1(){
  document.getElementById('step2').style.display='none';
  document.getElementById('step1').style.display='block';
}

async function doLogin(){
  var username=document.getElementById('l-user').value.trim();
  var pw=document.getElementById('l-pw').value;
  var err=document.getElementById('l-err');
  if(!username){err.style.display='block';err.textContent='أدخل اسم المستخدم';return;}
  sb('l-btn','جاري الدخول...',true);
  var r=await api('/api/login',{method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({username,pw,gid:selGid})});
  sb('l-btn','دخول',false);
  if(r.ok){
    err.style.display='none'; S={...r};
    if(r.role==='admin'){loadAdmin();show('s-admin');}
    else{document.getElementById('w-name').textContent=r.name;loadWorker();show('s-worker');}
  } else{err.style.display='block';err.textContent=r.msg||'خطأ';}
}

function doLogout(){
  api('/api/logout',{method:'POST'});
  S=null; CR=null;
  document.getElementById('step2').style.display='none';
  document.getElementById('step1').style.display='block';
  loadLoginGroups();
  show('s-login');
}

// ── Admin ───────────────────────────────────────────────────
var curTab='ov';
function aTab(name,el){
  curTab=name;
  document.querySelectorAll('.tab').forEach(b=>b.classList.remove('on'));
  el.classList.add('on');
  ['ov','logs','perf','emp','grp','bak'].forEach(t=>{
    var e=document.getElementById('t-'+t);
    if(e) e.style.display=(t===name?'block':'none');
  });
  if(name==='logs') loadLogs(1);
  else if(name==='perf') loadPerf('day');
  else if(name==='emp') loadEmp();
  else if(name==='grp') loadGrpMgr();
  else if(name==='bak') loadBak();
}

async function loadAdmin(){
  var r=await api('/api/stats');
  if(!r.ok) return;
  document.getElementById('at').textContent=r.total;
  document.getElementById('aa').textContent=r.active;
  document.getElementById('ad').textContent=r.departed;
  document.getElementById('ai').textContent=r.inactive;
  document.getElementById('adm-grps').innerHTML=r.groups.map(g=>`
    <div class="g-row">
      <div>
        <div class="g-name">${g.name}</div>
        <div class="g-sub">${g.total} حاج</div>
      </div>
      <div style="display:flex;gap:6px;flex-wrap:wrap;align-items:center">
        ${badge(g.active+' مفعّل','b-ok')}
        ${g.departed?badge(g.departed+' غادر','b-warn'):''}
        ${g.inactive?badge(g.inactive+' غير مفعّل','b-err'):''}
        <button class="nav-btn" onclick="viewGrp('${g.id}','${g.name}')">عرض</button>
      </div>
    </div>`).join('');
}

function viewGrp(gid,name){
  listGid=gid;
  document.querySelector('#s-list .nav-title').textContent='قائمة — '+name;
  show('s-list'); loadList(1);
}

async function doGSearch(){
  var q=document.getElementById('gs').value.trim(); if(!q) return;
  sb('gs-btn','جاري البحث...',true);
  var r=await api('/api/search?passport='+encodeURIComponent(q));
  sb('gs-btn','بحث في جميع المجموعات',false);
  var box=document.getElementById('gr');
  if(!r.ok){toast(r.msg,true);box.style.display='none';return;}
  document.getElementById('gr-g').textContent=r.group_name||r.group||'';
  document.getElementById('gr-nt').textContent=r.notes||'—';
  document.getElementById('gr-dat').textContent=r.departed_at||'—';
  document.getElementById('gr-st').innerHTML=r.status==='مفعل'?badge('مفعّل','b-ok'):badge('غير مفعّل','b-err');
  document.getElementById('gr-dp').innerHTML=r.departed==='نعم'?badge('غادر','b-warn'):badge('لم يغادر','b-blue');
  var seqEl=document.getElementById('gr-seq');
  if(seqEl) seqEl.textContent=r.seq_code||(r.seq?'م'+r.seq:'—');
  var diplBadge=document.getElementById('gr-dipl-badge');
  var diplBtn=document.getElementById('gr-dipl-btn');
  if(diplBadge){
    diplBadge.innerHTML=r.is_diplomatic?badge('دبلوماسي','b-dipl'):badge('عادي','b-gray');
    diplBtn.textContent=r.is_diplomatic?'إلغاء الدبلوماسي':'تعيين دبلوماسي';
    diplBtn.setAttribute('data-pp',r.passport);
    diplBtn.setAttribute('data-val',r.is_diplomatic?'0':'1');
  }
  box.style.display='block';
}
async function toggleDiplomatic(){
  var btn=document.getElementById('gr-dipl-btn');
  var pp=btn.getAttribute('data-pp');
  var val=parseInt(btn.getAttribute('data-val'));
  var r=await api('/api/update',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({passport:pp,is_diplomatic:val})});
  if(r.ok){
    toast(val?'تم تعيين الجواز دبلوماسي':'تم إلغاء الوضع الدبلوماسي');
    doGSearch();
  } else toast(r.msg||'فشل',true);
}

// ── Logs ────────────────────────────────────────────────────
async function loadLogs(page){
  logsPage=page;
  var u=document.getElementById('lf-u').value;
  var a=document.getElementById('lf-a').value;
  var p=document.getElementById('lf-p').value;
  var r=await api(`/api/logs?page=${page}&per=50&user=${encodeURIComponent(u)}&action=${encodeURIComponent(a)}&passport=${encodeURIComponent(p)}`);
  if(!r.ok) return;
  var body=document.getElementById('logs-body');
  if(!r.rows.length){body.innerHTML='<tr><td colspan="4" style="text-align:center;padding:20px;color:var(--tx3)">لا توجد نتائج</td></tr>';return;}
  var ac={'دخول':'b-blue','خروج':'b-gray','إضافة':'b-ok','تعديل':'b-blue','مغادرة':'b-warn','حذف':'b-err'};
  body.innerHTML=r.rows.map(row=>`<tr>
    <td style="font-size:12px;direction:ltr;white-space:nowrap">${row.ts||''}</td>
    <td>${row.username||''}</td>
    <td>${badge(row.action||'',ac[row.action]||'b-gray')}</td>
    <td style="direction:ltr;font-weight:600">${row.passport||''}</td>
  </tr>`).join('');
  var pages=Math.ceil(r.total/50);
  document.getElementById('logs-pag').innerHTML=pages>1?`
    <button onclick="loadLogs(${page-1})" ${page<=1?'disabled':''}>◀</button>
    <span>${page} / ${pages}</span>
    <button onclick="loadLogs(${page+1})" ${page>=pages?'disabled':''}>▶</button>`:'';
}

// ── Performance ─────────────────────────────────────────────
async function loadPerf(period,el){
  if(el){document.querySelectorAll('#t-perf .fb').forEach(b=>b.classList.remove('on'));el.classList.add('on');}
  var r=await api('/api/emp-stats?period='+period);
  if(!r.ok) return;
  var body=document.getElementById('perf-body');
  if(!r.stats.length){body.innerHTML='<tr><td colspan="8" style="text-align:center;padding:20px;color:var(--tx3)">لا توجد بيانات</td></tr>';return;}
  body.innerHTML=r.stats.map((s,i)=>{
    var tot=s.add+s.edit+s.depart+s.delete;
    return `<tr>
      <td style="color:var(--tx3)">#${i+1}</td>
      <td style="font-weight:600">${s.username}</td>
      <td style="color:var(--tx2)">${s.group||'—'}</td>
      <td style="color:var(--ok-tx);font-weight:600">${s.add}</td>
      <td style="color:var(--accent);font-weight:600">${s.edit}</td>
      <td style="color:var(--warn-tx);font-weight:600">${s.depart}</td>
      <td style="color:var(--err-tx);font-weight:600">${s.delete}</td>
      <td style="font-weight:700">${tot}</td>
    </tr>`;
  }).join('');
}

// ── Employees ───────────────────────────────────────────────
async function loadEmp(){
  var r=await api('/api/users');
  if(!r.ok) return;
  document.getElementById('emp-list').innerHTML=r.users.length?r.users.map(u=>`
    <div class="g-row">
      <div>
        <div class="g-name">${u.username}</div>
        <div class="g-sub">${u.role==='admin'?'مشرف':'موظف'}${u.group_name?' — '+u.group_name:''}</div>
      </div>
      <div style="display:flex;gap:6px;align-items:center">
        ${badge(u.active?'مفعّل':'معطّل',u.active?'b-ok':'b-err')}
        <button class="nav-btn" onclick='openUserModal(${JSON.stringify(u)})'>تعديل</button>
        <button class="nav-btn" style="color:var(--err-tx)" onclick="askDelUser(${u.id},'${u.username}')">حذف</button>
      </div>
    </div>`).join('')
    :'<div style="text-align:center;padding:20px;color:var(--tx3)">لا يوجد موظفون</div>';
}

function toggleGroupField(){
  var r=document.getElementById('um-role').value;
  document.getElementById('um-grp-f').style.display=r==='admin'?'none':'block';
}

async function openUserModal(user){
  var gr=await api('/api/groups');
  var opts=gr.ok?gr.groups.map(g=>`<option value="${g.id}">${g.name}</option>`).join(''):'';
  document.getElementById('um-grp').innerHTML='<option value="">— اختر —</option>'+opts;
  document.getElementById('um-err').style.display='none';
  document.getElementById('um-pw').value='';
  if(user){
    document.getElementById('um-title').textContent='تعديل موظف';
    document.getElementById('um-id').value=user.id;
    document.getElementById('um-user').value=user.username;
    document.getElementById('um-user').disabled=true;
    document.getElementById('um-role').value=user.role;
    document.getElementById('um-grp').value=user.group_id||'';
    document.getElementById('um-active').value=user.active?'1':'0';
  } else {
    document.getElementById('um-title').textContent='إضافة موظف';
    document.getElementById('um-id').value='';
    document.getElementById('um-user').value='';
    document.getElementById('um-user').disabled=false;
    document.getElementById('um-role').value='worker';
    document.getElementById('um-grp').value='';
    document.getElementById('um-active').value='1';
  }
  toggleGroupField();
  document.getElementById('m-user').classList.add('open');
}

async function saveUser(){
  var id=document.getElementById('um-id').value;
  var data={id:id||null,
    username:document.getElementById('um-user').value.trim(),
    password:document.getElementById('um-pw').value,
    role:document.getElementById('um-role').value,
    group_id:document.getElementById('um-grp').value,
    active:parseInt(document.getElementById('um-active').value)};
  var r=await api('/api/users/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});
  if(r.ok){close('m-user');toast('تم الحفظ ✓');loadEmp();}
  else{var e=document.getElementById('um-err');e.style.display='block';e.textContent=r.msg;}
}

var _delUserId=null;
function askDelUser(id,name){
  if(confirm(`حذف المستخدم "${name}"؟`))delUser(id);
}
async function delUser(id){
  var r=await api('/api/users/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});
  if(r.ok){toast('تم الحذف');loadEmp();}
  else toast(r.msg,true);
}

// ── Groups Mgmt ─────────────────────────────────────────────
async function loadGrpMgr(){
  var r=await api('/api/groups');
  if(!r.ok) return;
  document.getElementById('grp-list').innerHTML=r.groups.length?r.groups.map(g=>`
    <div class="g-row">
      <div>
        <div class="g-name" style="display:flex;align-items:center;gap:8px">
          ${g.name}
          <span style="background:var(--accent);color:#fff;border-radius:6px;padding:1px 8px;
            font-size:12px;font-weight:700;direction:ltr;letter-spacing:1px">${g.prefix||'?'}</span>
        </div>
        <div class="g-sub" style="direction:ltr">${g.id} — ترقيم: ${g.prefix||'?'}-1، ${g.prefix||'?'}-2...</div>
      </div>
      <div style="display:flex;gap:6px">
        <button class="nav-btn" onclick='openGroupModal(${JSON.stringify(g)})'>تعديل</button>
        <button class="nav-btn" style="color:var(--err-tx)" onclick="askDelGrp('${g.id}','${g.name}')">حذف</button>
      </div>
    </div>`).join('')
    :'<div style="text-align:center;padding:20px;color:var(--tx3)">لا توجد مجموعات</div>';
}

function openGroupModal(grp){
  document.getElementById('gm-err').style.display='none';
  var pfxInp=document.getElementById('gm-prefix');
  if(grp){
    document.getElementById('gm-title').textContent='تعديل مجموعة';
    document.getElementById('gm-orig').value=grp.id;
    document.getElementById('gm-id').value=grp.id;
    document.getElementById('gm-id').disabled=true;
    document.getElementById('gm-name').value=grp.name;
    pfxInp.value=grp.prefix||'';
  } else {
    document.getElementById('gm-title').textContent='إضافة مجموعة';
    document.getElementById('gm-orig').value='';
    document.getElementById('gm-id').value='';
    document.getElementById('gm-id').disabled=false;
    document.getElementById('gm-name').value='';
    pfxInp.value='';
  }
  updatePrefixPreview();
  pfxInp.oninput=function(){this.value=this.value.toUpperCase();updatePrefixPreview();};
  document.getElementById('m-grp').classList.add('open');
}

function updatePrefixPreview(){
  var p=document.getElementById('gm-prefix').value.toUpperCase()||'?';
  document.getElementById('gm-preview').textContent=p+'-1، '+p+'-2...';
}

async function saveGroup(){
  var data={id:document.getElementById('gm-id').value,
    name:document.getElementById('gm-name').value,
    prefix:document.getElementById('gm-prefix').value.toUpperCase(),
    orig_id:document.getElementById('gm-orig').value};
  var r=await api('/api/groups/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});
  if(r.ok){close('m-grp');toast('تم الحفظ ✓');loadGrpMgr();loadAdmin();}
  else{var e=document.getElementById('gm-err');e.style.display='block';e.textContent=r.msg;}
}

function askDelGrp(id,name){
  document.getElementById('delgrp-id').value=id;
  document.getElementById('delgrp-name').textContent=name;
  document.getElementById('m-delgrp').classList.add('open');
}
async function confirmDelGrp(){
  var id=document.getElementById('delgrp-id').value;
  var act=document.getElementById('delgrp-emp').value;
  var r=await api('/api/groups/delete',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({id,emp_action:act})});
  close('m-delgrp');
  if(r.ok){toast('تم حذف المجموعة');loadGrpMgr();loadAdmin();}
  else toast(r.msg,true);
}

// ── Backup ──────────────────────────────────────────────────
async function loadBak(){
  var r=await api('/api/backups');
  if(!r.ok) return;
  var el=document.getElementById('bak-list');
  el.innerHTML=r.files.length?r.files.map(f=>`
    <div class="row"><span class="row-v" style="font-size:13px;direction:ltr">${f}</span>
    ${badge('✓','b-ok')}</div>`).join('')
    :'<div style="text-align:center;padding:16px;color:var(--tx3)">لا توجد نسخ</div>';
}
async function doBak(){
  sb('bak-btn','جاري النسخ...',true);
  var r=await api('/api/backup',{method:'POST'});
  sb('bak-btn','💾 إنشاء نسخة الآن',false);
  if(r.ok){toast('تم: '+r.file);loadBak();}
  else toast(r.msg,true);
}

// ── Worker ──────────────────────────────────────────────────
async function loadWorker(){
  var r=await api('/api/stats?gid='+S.gid);
  if(!r.ok) return;
  document.getElementById('wt').textContent=r.total||0;
  document.getElementById('wa').textContent=r.active||0;
  document.getElementById('wd').textContent=r.departed||0;
  document.getElementById('wi').textContent=r.inactive||0;
}

async function doWSearch(){
  var q=document.getElementById('ws').value.trim(); if(!q) return;
  sb('ws-btn','جاري البحث...',true);
  var r=await api('/api/search?passport='+encodeURIComponent(q));
  sb('ws-btn','بحث',false);
  var box=document.getElementById('wr');
  if(!r.ok){toast(r.msg,true);box.style.display='none';return;}
  CR=r;
  document.getElementById('wr-pp').textContent=r.passport;
  document.getElementById('wr-seq').textContent=r.seq_code||(r.seq?'م'+r.seq:'—');
  document.getElementById('wr-nt').textContent=r.notes||'—';
  var diplEl=document.getElementById('wr-dipl');
  if(r.is_diplomatic){diplEl.innerHTML=badge('جواز دبلوماسي','b-dipl');diplEl.style.display='block';}
  else{diplEl.innerHTML='';diplEl.style.display='none';}
  document.getElementById('w-notes').value=r.notes||'';
  document.getElementById('t-st').checked=r.status==='مفعل';
  var tag=document.getElementById('wr-tag');
  tag.textContent=r.status==='مفعل'?'مفعّل':'غير مفعّل';
  tag.className='badge '+(r.status==='مفعل'?'b-ok':'b-err');
  box.style.display='block';
}

async function doSave(){
  if(!CR) return;
  var st=document.getElementById('t-st').checked?'مفعل':'غير مفعل';
  var nt=document.getElementById('w-notes').value;
  sb('sv-btn','جاري الحفظ...',true);
  var r=await api('/api/update',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({passport:CR.passport,status:st,notes:nt})});
  sb('sv-btn','حفظ',false);
  if(r.ok){
    toast('تم الحفظ ✓'); loadWorker();
    CR.status=st; CR.notes=nt;
    document.getElementById('wr-nt').textContent=nt||'—';
    var tag=document.getElementById('wr-tag');
    tag.textContent=st==='مفعل'?'مفعّل':'غير مفعّل';
    tag.className='badge '+(st==='مفعل'?'b-ok':'b-err');
  } else toast(r.msg,true);
}

function doDelete(){
  if(!CR) return;
  document.getElementById('del-pp').textContent=CR.passport;
  document.getElementById('m-del').classList.add('open');
}
async function confirmDel(){
  close('m-del');
  var r=await api('/api/delete',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({passport:CR.passport})});
  if(r.ok){
    toast('تم الحذف'); CR=null;
    document.getElementById('wr').style.display='none';
    document.getElementById('ws').value='';
    loadWorker();
  } else toast(r.msg,true);
}

// ── Add ─────────────────────────────────────────────────────
async function doAdd(){
  var ppInp=document.getElementById('a-pp');
  var pp=ppInp.value.trim().toUpperCase();
  var st=document.getElementById('a-st').value;
  var nt=document.getElementById('a-nt').value;
  var err=document.getElementById('a-err'), ok=document.getElementById('a-ok');
  err.style.display='none'; ok.style.display='none';
  ppInp.style.borderColor='';
  if(!pp){
    err.style.display='block';err.textContent='أدخل رقم الجواز';
    ppInp.style.borderColor='var(--err)';ppInp.focus();return;
  }
  if(!/^[A-Z]{2}[0-9]{7}$/.test(pp)){
    err.style.display='block';err.textContent='صيغة الجواز غير صحيحة — يجب حرفان + 7 أرقام (مثال: AB1234567)';
    ppInp.style.borderColor='var(--err)';ppInp.focus();return;
  }
  if(!st){err.style.display='block';err.textContent='اختر حالة النسك';return;}
  sb('a-btn','جاري الإضافة...',true);
  var r=await api('/api/add',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({gid:S.gid,passport:pp,status:st,notes:nt})});
  sb('a-btn','➕ إضافة',false);
  if(r.ok){
    var diplNote = r.is_diplomatic ? ' — 🔴 دبلوماسي' : '';
    ok.style.display='block'; ok.textContent='تمت الإضافة ✓  الرقم: '+(r.seq_code||r.seq)+diplNote;
    ppInp.value=''; ppInp.style.borderColor='';
    document.getElementById('a-st').value='';
    document.getElementById('a-nt').value='';
    if(S.role!=='admin') loadWorker();
  } else{
    err.style.display='block';err.textContent=r.msg||'فشل';
    ppInp.style.borderColor='var(--err)';ppInp.focus();
  }
}

// ── List ────────────────────────────────────────────────────
function showList(gid){
  listGid=gid||S.gid||'';
  if(!gid) document.querySelector('#s-list .nav-title').textContent='قائمة الحجاج';
  show('s-list'); loadList(1);
}
async function loadList(page){
  listPage=page||1;
  document.getElementById('list-rows').innerHTML='<div style="text-align:center;padding:24px;color:var(--tx3)">جاري التحميل...</div>';
  var r=await api(`/api/list?gid=${listGid}&filter=${curF}&page=${listPage}&per=50`);
  if(!r.ok){toast(r.msg,true);return;}
  listData=r.rows;
  renderList(listData);
  document.getElementById('list-cnt').textContent=r.total+' سجل';
  var pages=Math.ceil(r.total/50);
  document.getElementById('list-pag').innerHTML=pages>1?`
    <button onclick="loadList(${listPage-1})" ${listPage<=1?'disabled':''}>◀</button>
    <span>${listPage} / ${pages}</span>
    <button onclick="loadList(${listPage+1})" ${listPage>=pages?'disabled':''}>▶</button>`:'';
}
function setF(f,el){
  curF=f;
  document.querySelectorAll('#s-list .fb').forEach(b=>b.classList.remove('on'));
  el.classList.add('on'); loadList(1);
}
function fLocal(){
  var q=document.getElementById('list-q').value.trim().toUpperCase();
  if(!q){renderList(listData);return;}
  renderList(listData.filter(r=>r.passport.toUpperCase().includes(q)||(r.notes||'').toUpperCase().includes(q)));
}
function renderList(rows){
  if(!rows.length){document.getElementById('list-rows').innerHTML='<div style="text-align:center;padding:24px;color:var(--tx3)">لا توجد نتائج</div>';return;}
  document.getElementById('list-rows').innerHTML=rows.map(r=>{
    var st=r.status==='مفعل'?badge('مفعّل','b-ok'):badge('غير مفعّل','b-err');
    var dp=r.departed==='نعم'?badge('غادر','b-warn'):badge('لم يغادر','b-gray');
    var code=r.seq_code||('م'+r.seq);
    var diplBadge=r.is_diplomatic?badge('دبلوماسي','b-dipl'):'';
    var rowCls=r.is_diplomatic?' dipl':'';
    return `<div class="list-row${rowCls}" onclick="openFromList('${r.passport}')">
      <div>
        <div class="lp" style="display:flex;align-items:center;gap:6px">${r.passport}${diplBadge}</div>
        <div class="ln" style="display:flex;gap:8px;align-items:center">
          <span style="font-weight:700;color:var(--accent);direction:ltr">${code}</span>
          ${r.notes?'<span>— '+r.notes.substring(0,20)+'</span>':''}
        </div>
      </div>
      <div class="badge-row">${st}${dp}</div></div>`;
  }).join('');
}
function openFromList(passport){
  if(S.role==='admin'){
    document.getElementById('gs').value=passport;
    show('s-admin'); doGSearch();
  } else {
    document.getElementById('ws').value=passport;
    show('s-worker'); doWSearch();
  }
}

// ── Depart ──────────────────────────────────────────────────
function showDepart(){
  show('s-depart');
  document.getElementById('dp-q').value='';
  document.getElementById('dp-res').style.display='none';
}
async function doDSearch(){
  var q=document.getElementById('dp-q').value.trim(); if(!q) return;
  sb('dp-btn','جاري البحث...',true);
  var r=await api('/api/search?passport='+encodeURIComponent(q));
  sb('dp-btn','🔍 بحث',false);
  var box=document.getElementById('dp-res');
  if(!r.ok){toast(r.msg,true);box.style.display='none';return;}
  document.getElementById('dp-pp').textContent=r.passport;
  document.getElementById('dp-grp').textContent=r.group_name||r.group||'';
  document.getElementById('dp-nt').textContent=r.notes||'—';
  var dpDipl=document.getElementById('dp-dipl');
  if(dpDipl){if(r.is_diplomatic){dpDipl.innerHTML=badge('جواز دبلوماسي','b-dipl');dpDipl.style.display='block';}
  else{dpDipl.innerHTML='';dpDipl.style.display='none';}}
  document.getElementById('dp-st').innerHTML=r.status==='مفعل'?badge('مفعّل','b-ok'):badge('غير مفعّل','b-err');
  var timeRow=document.getElementById('dp-time-r');
  var action=document.getElementById('dp-action');
  var tag=document.getElementById('dp-tag');
  if(r.departed==='نعم'){
    tag.textContent='غادر ✓'; tag.className='badge b-warn';
    timeRow.style.display='flex';
    document.getElementById('dp-time').textContent=r.departed_at||'—';
    action.innerHTML='<div style="text-align:center;padding:12px;color:var(--warn-tx);font-weight:600">✈️ غادر مسبقاً</div>';
    box.className='card dep-done';
  } else {
    tag.textContent='لم يغادر'; tag.className='badge b-gray';
    timeRow.style.display='none';
    action.innerHTML='<button class="btn btn-warn" id="dp-cfm" onclick="doCDepart()">✈️ تأكيد المغادرة</button>';
    box.className='card dep-card';
  }
  box.setAttribute('data-pp',r.passport);
  box.style.display='block';
}
async function doCDepart(){
  var pp=document.getElementById('dp-res').getAttribute('data-pp'); if(!pp) return;
  sb('dp-cfm','جاري التأكيد...',true);
  var r=await api('/api/depart',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({passport:pp})});
  if(r.ok){
    toast('تم تسجيل المغادرة ✓');
    document.getElementById('dp-tag').textContent='غادر ✓';
    document.getElementById('dp-tag').className='badge b-warn';
    document.getElementById('dp-time-r').style.display='flex';
    document.getElementById('dp-time').textContent=r.departed_at;
    document.getElementById('dp-action').innerHTML='<div style="text-align:center;padding:12px;color:var(--warn-tx);font-weight:600">✈️ تم تسجيل المغادرة</div>';
    document.getElementById('dp-res').className='card dep-done';
    if(S.role!=='admin') loadWorker();
  } else toast(r.msg,true);
}

function exportExcel(){window.location.href='/api/export';toast('جاري تحميل الملف...');}

// ── Init ────────────────────────────────────────────────────
(async function(){
  loadLoginGroups();
  var r=await api('/api/session');
  if(r.ok){
    S={role:r.role,name:r.name,gid:r.gid};
    if(r.role==='admin'){loadAdmin();show('s-admin');}
    else{document.getElementById('w-name').textContent=r.name;loadWorker();show('s-worker');}
  }
})();
</script>
</body>
</html>"""

if __name__ == "__main__":
    init_db()
    migrate_json()
    migrate_seq_codes()
    os.makedirs(BACKUP_DIR, exist_ok=True)
    threading.Thread(target=backup_scheduler, daemon=True).start()

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80)); ip = s.getsockname()[0]
    except:
        ip = "127.0.0.1"
    finally:
        s.close()

    print("=" * 52)
    print("🕋  نظام تنظيم الجوازات — v1.4")
    print("=" * 52)
    print(f"\n✅ السيرفر شغّال!")
    print(f"\n🔗 جهازك:    http://localhost:5000")
    print(f"🔗 الشبكة:   http://{ip}:5000")
    print("\n📋 بيانات الدخول:")
    print("   مشرف:  admin / admin2024")
    print("   موظف:  fagh  / 1234  (مثال)")
    print("\n⚠️  لا تغلق هذه النافذة أثناء العمل")
    print("=" * 52)
    if __name__ == "__main__":
    init_db()
    migrate_json()
    migrate_seq_codes()
    os.makedirs(BACKUP_DIR, exist_ok=True)
    threading.Thread(target=backup_scheduler, daemon=True).start()

    app.run(host="0.0.0.0", port=5000)
