import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from flask import Flask, request, jsonify, send_file, session, redirect, render_template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
from xml.sax.saxutils import escape
from copy import deepcopy
import xml.etree.ElementTree as ET
import os, io, sqlite3, json, shutil, threading, time, socket, re, zipfile

app = Flask(__name__)
app.secret_key = b'\x9f\x4a\x2c\x8e\x1b\x7d\x3f\x56\xa0\xc2\x4e\x91\x68\x0b\xd5\x37'
app.permanent_session_lifetime = timedelta(hours=8)

DB_FILE     = "passports.db"
BACKUP_DIR  = "backups"
UPLOADS_DIR = "uploads"
TRIPS_DAILY_TEMPLATE = os.path.join("report_templates", "trips_daily_template.docx")
MAX_BACKUPS = 7

# ── DB ───────────────────────────────────────────────────────

def get_db():
    db = sqlite3.connect(DB_FILE, check_same_thread=False, timeout=10)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    return db

DEFAULT_PREFIXES = {
    "fagh":"A", "amsa":"B", "cuba":"C",
    "safar":"D", "top":"E", "top10":"F"
}

def init_db():
    db = get_db()
    try:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS groups (
                id     TEXT PRIMARY KEY,
                name   TEXT NOT NULL,
                prefix TEXT,
                icon   TEXT DEFAULT 'users',
                color  TEXT DEFAULT '#3B82F6'
            );
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                username      TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role          TEXT NOT NULL DEFAULT 'worker',
                group_id      TEXT,
                active        INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS pilgrims (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                passport       TEXT UNIQUE NOT NULL,
                group_id       TEXT NOT NULL,
                seq            INTEGER,
                seq_code       TEXT,
                status         TEXT DEFAULT '',
                is_diplomatic  INTEGER DEFAULT 0,
                departed       TEXT DEFAULT 'لا',
                departed_at    TEXT,
                notes          TEXT DEFAULT '',
                deleted        INTEGER NOT NULL DEFAULT 0,
                created_at     TEXT,
                updated_at     TEXT
            );
            CREATE TABLE IF NOT EXISTS audit_logs (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                username   TEXT,
                user_group TEXT,
                action     TEXT,
                passport   TEXT,
                old_values TEXT,
                new_values TEXT,
                ts         TEXT
            );
            CREATE TABLE IF NOT EXISTS attachments (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                passport   TEXT NOT NULL,
                filename   TEXT NOT NULL,
                mime       TEXT NOT NULL DEFAULT 'application/octet-stream',
                data       TEXT NOT NULL DEFAULT '',
                filepath   TEXT,
                created_at TEXT
            );
            CREATE TABLE IF NOT EXISTS trips (
                id TEXT PRIMARY KEY,
                group_id TEXT,
                hotel_name TEXT,
                location_url TEXT,
                nationality TEXT,
                flight_no TEXT,
                carrier TEXT,
                destination TEXT,
                approval_no TEXT,
                departure_time TEXT,
                housing_contract_no TEXT,
                bus_stand_time TEXT,
                bus_departure_time TEXT,
                bus_count INTEGER,
                planned_pilgrim_count INTEGER,
                report_date TEXT,
                notes TEXT,
                status TEXT NOT NULL DEFAULT 'draft',
                created_at TEXT
            );
        """)
        for stmt in ["ALTER TABLE groups       ADD COLUMN prefix        TEXT",
                     "ALTER TABLE groups       ADD COLUMN icon         TEXT DEFAULT 'users'",
                     "ALTER TABLE groups       ADD COLUMN color        TEXT DEFAULT '#3B82F6'",
                     "ALTER TABLE pilgrims     ADD COLUMN seq_code     TEXT",
                     "ALTER TABLE pilgrims     ADD COLUMN is_diplomatic INTEGER DEFAULT 0",
                     "ALTER TABLE pilgrims     ADD COLUMN deleted       INTEGER NOT NULL DEFAULT 0",
                     "ALTER TABLE pilgrims     ADD COLUMN depart_batch  TEXT",
                     "ALTER TABLE pilgrims     ADD COLUMN trip_id       TEXT",
                     "ALTER TABLE trips        ADD COLUMN group_id      TEXT",
                     "ALTER TABLE trips        ADD COLUMN location_url  TEXT",
                     "ALTER TABLE trips        ADD COLUMN nationality   TEXT",
                     "ALTER TABLE trips        ADD COLUMN approval_no   TEXT",
                     "ALTER TABLE trips        ADD COLUMN housing_contract_no TEXT",
                     "ALTER TABLE trips        ADD COLUMN bus_stand_time TEXT",
                     "ALTER TABLE trips        ADD COLUMN bus_departure_time TEXT",
                     "ALTER TABLE trips        ADD COLUMN planned_pilgrim_count INTEGER",
                     "ALTER TABLE trips        ADD COLUMN report_date   TEXT",
                     "ALTER TABLE trips        ADD COLUMN notes         TEXT",
                     "ALTER TABLE trips        ADD COLUMN status        TEXT NOT NULL DEFAULT 'draft'",
                     "ALTER TABLE attachments  ADD COLUMN filepath     TEXT"]:
            try:
                db.execute(stmt)
            except Exception:
                pass  # العمود موجود مسبقاً
        db.executescript("""
            CREATE INDEX IF NOT EXISTS idx_pilgrims_passport ON pilgrims(passport);
            CREATE INDEX IF NOT EXISTS idx_pilgrims_group    ON pilgrims(group_id);
            CREATE INDEX IF NOT EXISTS idx_pilgrims_deleted  ON pilgrims(deleted);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_group_seq   ON pilgrims(group_id, seq);
            CREATE INDEX IF NOT EXISTS idx_audit_ts          ON audit_logs(id);
        """)
        if not db.execute("SELECT 1 FROM users WHERE username='admin'").fetchone():
            db.execute("INSERT INTO users (username,password_hash,role) VALUES (?,?,?)",
                       ("admin", generate_password_hash("admin2024"), "admin"))
        db.commit()
    finally:
        db.close()

def _group_prefix(db, group_id):
    row = db.execute("SELECT prefix FROM groups WHERE id=?", (group_id,)).fetchone()
    if row and row["prefix"]:
        return row["prefix"]
    return DEFAULT_PREFIXES.get(group_id, group_id[0].upper() if group_id else "X")

def next_seq_code(db, group_id):
    """Return (seq_int, seq_code) for next record in group — never reuses after delete."""
    max_seq = db.execute(
        "SELECT COALESCE(MAX(seq),0) as m FROM pilgrims WHERE group_id=?", (group_id,)
    ).fetchone()["m"]
    seq = max_seq + 1
    prefix = _group_prefix(db, group_id)
    return seq, f"{prefix}-{seq}"

def migrate_seq_codes():
    db = get_db()
    try:
        rows = db.execute("""
            SELECT p.id, p.seq, p.group_id
            FROM pilgrims p
            WHERE p.seq_code IS NULL OR p.seq_code = ''
        """).fetchall()
        for r in rows:
            prefix = _group_prefix(db, r["group_id"])
            code   = f"{prefix}-{r['seq'] or 0}"
            db.execute("UPDATE pilgrims SET seq_code=? WHERE id=?", (code, r["id"]))
        if rows:
            db.commit()
            print(f"✅ تم إنشاء seq_code لـ {len(rows)} سجل")
    finally:
        db.close()

def migrate_json():
    if not os.path.exists("data.json") or os.path.exists("data.json.migrated"):
        return
    try:
        with open("data.json","r",encoding="utf-8") as f:
            data = json.load(f)
        db = get_db()
        try:
            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            for gid, rows in data.items():
                for r in rows:
                    db.execute("""INSERT OR IGNORE INTO pilgrims
                        (passport,group_id,seq,status,departed,notes,created_at,updated_at)
                        VALUES (?,?,?,?,?,?,?,?)""",
                        (r.get("passport","").upper(), gid, r.get("seq"),
                         r.get("status",""), r.get("departed","لا"),
                         r.get("notes",""), r.get("updated",now), r.get("updated",now)))
            db.commit()
            open("data.json.migrated","w").close()
            print("✅ تم نقل data.json إلى SQLite")
        finally:
            db.close()
    except Exception as e:
        print(f"⚠️ خطأ في النقل: {e}")

# ── Helpers ──────────────────────────────────────────────────

def log_action(action, passport=None, old=None, new=None):
    db = None
    try:
        db = get_db()
        db.execute("INSERT INTO audit_logs (username,user_group,action,passport,old_values,new_values,ts) VALUES (?,?,?,?,?,?,?)",
            (session.get("username","?"), session.get("group_id",""), action, passport,
             json.dumps(old, ensure_ascii=False) if old else None,
             json.dumps(new, ensure_ascii=False) if new else None,
             datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        db.commit()
    except Exception as e:
        print(f"ERROR log_action: {e}")
    finally:
        if db: db.close()

def require_login(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("user_id"):
            return jsonify({"ok":False,"msg":"غير مسجّل الدخول","unauth":True}), 401
        return f(*a,**k)
    return d

def require_admin(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("user_id") or session.get("role") != "admin":
            return jsonify({"ok":False,"msg":"غير مصرح"}), 403
        return f(*a,**k)
    return d

def migrate_attachments_to_fs():
    import base64
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    db = get_db()
    rows = db.execute(
        "SELECT id, filename, data FROM attachments WHERE (filepath IS NULL OR filepath='') AND data != ''"
    ).fetchall()
    migrated = 0
    for row in rows:
        try:
            _, ext = os.path.splitext(row["filename"])
            ext = re.sub(r'[^a-zA-Z0-9.]', '', ext)[:10]
            fpath = os.path.join(UPLOADS_DIR, f"{row['id']}{ext}")
            with open(fpath, 'wb') as f:
                f.write(base64.b64decode(row["data"]))
            db.execute("UPDATE attachments SET filepath=?, data='' WHERE id=?", (fpath, row["id"]))
            migrated += 1
        except Exception as e:
            print(f"WARN migrate attachment {row['id']}: {e}")
    if migrated:
        db.commit()
        print(f"✓ نُقل {migrated} مرفق من قاعدة البيانات إلى القرص")
    db.close()

def do_backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_DIR, f"passports_{ts}.db")
    shutil.copy2(DB_FILE, dst)
    # نسخ احتياطي لمجلد المرفقات
    if os.path.isdir(UPLOADS_DIR):
        uploads_bak = os.path.join(BACKUP_DIR, f"uploads_{ts}")
        shutil.copytree(UPLOADS_DIR, uploads_bak)
    # حذف النسخ القديمة (DB + مجلد uploads المقابل)
    db_files = sorted(f for f in os.listdir(BACKUP_DIR) if f.startswith("passports_") and f.endswith(".db"))
    for old in db_files[:-MAX_BACKUPS]:
        os.remove(os.path.join(BACKUP_DIR, old))
        old_ts  = old.replace("passports_","").replace(".db","")
        old_upl = os.path.join(BACKUP_DIR, f"uploads_{old_ts}")
        if os.path.isdir(old_upl):
            shutil.rmtree(old_upl)
    return dst

def backup_scheduler():
    while True:
        time.sleep(6*3600)
        try:
            do_backup()
        except Exception as e:
            print(f"ERROR backup_scheduler: {e}")

# ── Auth ─────────────────────────────────────────────────────

@app.route("/api/public/groups")
def public_groups():
    db = get_db()
    try:
        groups = db.execute("SELECT id,name,icon,color FROM groups ORDER BY name").fetchall()
        return jsonify({"ok":True,"groups":[dict(g) for g in groups]})
    finally:
        db.close()

@app.route("/api/login", methods=["POST"])
def login():
    d = request.json or {}
    username = d.get("username","").strip()
    pw       = d.get("pw","")
    gid      = d.get("gid","")
    if not username or not pw:
        return jsonify({"ok":False,"msg":"أدخل اسم المستخدم وكلمة السر"})
    db = get_db()
    try:
        user = db.execute("SELECT * FROM users WHERE username=? AND active=1",(username,)).fetchone()
        if not user or not check_password_hash(user["password_hash"], pw):
            return jsonify({"ok":False,"msg":"اسم المستخدم أو كلمة السر غلط"})
        if gid == "admin":
            if user["role"] != "admin":
                return jsonify({"ok":False,"msg":"هذا المستخدم ليس مشرفاً"})
        else:
            if user["role"] == "admin":
                return jsonify({"ok":False,"msg":"استخدم خيار المشرف للدخول"})
            if user["group_id"] != gid:
                return jsonify({"ok":False,"msg":"هذا المستخدم لا ينتمي لهذه المجموعة"})
        grp = db.execute("SELECT name FROM groups WHERE id=?", (user["group_id"],)).fetchone() if user["group_id"] else None
        session.clear()
        session.permanent = True
        session["user_id"]  = user["id"]
        session["username"] = user["username"]
        session["role"]     = user["role"]
        session["group_id"] = user["group_id"] or ""
        session["name"]     = "مشرف عام" if user["role"]=="admin" else (grp["name"] if grp else username)
        log_action("دخول")
        return jsonify({"ok":True,"role":user["role"],"name":session["name"],"gid":session["group_id"],
                        "username":session["username"]})
    finally:
        db.close()

@app.route("/api/logout", methods=["POST"])
def api_logout():
    log_action("خروج")
    session.clear()
    return jsonify({"ok":True})

@app.route("/api/session")
def check_session():
    if session.get("user_id"):
        return jsonify({"ok":True,"role":session.get("role"),"name":session.get("name"),"gid":session.get("group_id",""),
                        "username":session.get("username","")})
    return jsonify({"ok":False})

# ── Stats ────────────────────────────────────────────────────

@app.route("/api/stats")
@require_login
def stats():
    gid = request.args.get("gid","")
    if session["role"] != "admin":
        gid = session["group_id"]
    db = get_db()
    try:
        if gid and gid != "all":
            grp = db.execute("SELECT id,name FROM groups WHERE id=?", (gid,)).fetchone()
            if not grp:
                return jsonify({"ok":False,"msg":"المجموعة غير موجودة"}), 404
            rows = db.execute("SELECT status,departed FROM pilgrims WHERE group_id=? AND deleted=0",(gid,)).fetchall()
            a  = sum(1 for r in rows if r["status"]=="مفعل")
            i  = sum(1 for r in rows if r["status"]=="غير مفعل")
            dp = sum(1 for r in rows if r["departed"]=="نعم")
            pn = len(rows) - dp
            progress = int(dp/len(rows)*100) if rows else 0
            return jsonify({"ok":True,"total":len(rows),
                "active":a,"inactive":i,"departed":dp,"pending":pn,
                "group_id":grp["id"],"group_name":grp["name"],"progress":progress})
        groups = db.execute("SELECT * FROM groups").fetchall()
        result=[]; total=active=inactive=departed=pending=0
        for g in groups:
            rows = db.execute("SELECT status,departed FROM pilgrims WHERE group_id=? AND deleted=0",(g["id"],)).fetchall()
            t=len(rows); a=sum(1 for r in rows if r["status"]=="مفعل")
            i=sum(1 for r in rows if r["status"]=="غير مفعل")
            dp=sum(1 for r in rows if r["departed"]=="نعم")
            pn=t-dp
            total+=t; active+=a; inactive+=i; departed+=dp; pending+=pn
            result.append({"id":g["id"],"name":g["name"],"total":t,"active":a,"inactive":i,"departed":dp,
                           "pending":pn,"progress":(int(dp/t*100) if t else 0)})
        deleted_count = db.execute("SELECT COUNT(*) as c FROM pilgrims WHERE deleted=1").fetchone()["c"]
        progress = int(departed/total*100) if total else 0
        return jsonify({"ok":True,"total":total,"active":active,"inactive":inactive,
                        "departed":departed,"pending":pending,"deleted":deleted_count,
                        "groups":result,"progress":progress})
    finally:
        db.close()

# ── Search ───────────────────────────────────────────────────

@app.route("/api/search")
@require_login
def search():
    passport = request.args.get("passport","").strip().upper()
    if not passport:
        return jsonify({"ok":False,"msg":"أدخل رقم الجواز"})
    db = get_db()
    try:
        if session["role"] == "admin":
            row = db.execute("""SELECT p.*,g.name as group_name FROM pilgrims p
                JOIN groups g ON p.group_id=g.id WHERE UPPER(p.passport)=?""",(passport,)).fetchone()
            if not row: return jsonify({"ok":False,"msg":"غير موجود في أي مجموعة"})
            return jsonify({"ok":True,**dict(row)})
        row = db.execute("""SELECT p.*,g.name as group_name FROM pilgrims p
            JOIN groups g ON p.group_id=g.id WHERE UPPER(p.passport)=? AND p.deleted=0""",(passport,)).fetchone()
        if not row: return jsonify({"ok":False,"msg":"رقم الجواز غير موجود"})
        can_act = row["group_id"] == session["group_id"]
        return jsonify({"ok":True,"can_act":can_act,**dict(row)})
    finally:
        db.close()

# ── Add ──────────────────────────────────────────────────────

@app.route("/api/add", methods=["POST"])
@require_login
def add():
    d  = request.json or {}
    gid = session["group_id"] if session["role"]!="admin" else d.get("gid","")
    if not gid: return jsonify({"ok":False,"msg":"حدد المجموعة"})
    passport = d.get("passport","").strip().upper()
    if not passport: return jsonify({"ok":False,"msg":"أدخل رقم الجواز"})
    if not re.match(r'^[A-Z]{2}[0-9]{7}$', passport):
        return jsonify({"ok":False,"msg":"صيغة الجواز غير صحيحة — يجب حرفان + 7 أرقام (مثال: AB1234567)"})
    db = get_db()
    try:
        db.execute("BEGIN IMMEDIATE")
        ex = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
        if ex:
            if ex["deleted"]:
                now = datetime.now().strftime("%Y-%m-%d %H:%M")
                is_diplomatic = 1 if passport.startswith("DA") else 0
                status = d.get("status","غير مفعل") or "غير مفعل"; notes = d.get("notes","")
                db.execute("""UPDATE pilgrims SET deleted=0,group_id=?,status=?,is_diplomatic=?,
                              departed='لا',departed_at=NULL,notes=?,updated_at=? WHERE id=?""",
                           (gid,status,is_diplomatic,notes,now,ex["id"]))
                db.commit()
                log_action("استعادة", passport, new={"status":status,"group":gid})
                return jsonify({"ok":True,"seq":ex["seq"],"seq_code":ex["seq_code"],
                                "is_diplomatic":is_diplomatic,"restored":True})
            grp = db.execute("SELECT name FROM groups WHERE id=?",(ex["group_id"],)).fetchone()
            grp_name = grp['name'] if grp else ex['group_id']
            seq_info = f" — رقم {ex['seq_code']}" if ex['seq_code'] else (f" — رقم {ex['seq']}" if ex['seq'] else "")
            db.rollback()
            return jsonify({"ok":False,"msg":f"⚠️ مكرر — موجود في {grp_name}{seq_info}"})
        seq = db.execute(
            "SELECT COALESCE(MAX(seq),0)+1 AS seq FROM pilgrims WHERE group_id=?",
            (gid,)
        ).fetchone()["seq"]
        seq_code = f"{_group_prefix(db, gid)}-{seq}"
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        status = d.get("status","غير مفعل") or "غير مفعل"; notes = d.get("notes","")
        is_diplomatic = 1 if passport.startswith("DA") else 0
        db.execute("""INSERT INTO pilgrims
            (passport,group_id,seq,seq_code,status,is_diplomatic,departed,notes,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
                   (passport,gid,seq,seq_code,status,is_diplomatic,"لا",notes,now,now))
        db.commit()
        log_action("إضافة", passport, new={"status":status,"seq_code":seq_code,"is_diplomatic":is_diplomatic})
        return jsonify({"ok":True,"seq":seq,"seq_code":seq_code,"is_diplomatic":is_diplomatic})
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

# ── Update ───────────────────────────────────────────────────

@app.route("/api/update", methods=["POST"])
@require_login
def update():
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    db = get_db()
    try:
        row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
        if not row: return jsonify({"ok":False,"msg":"السجل غير موجود"})
        if session["role"]!="admin" and row["group_id"]!=session["group_id"]:
            return jsonify({"ok":False,"msg":"غير مصرح"})
        old  = dict(row)
        nst  = d.get("status",  row["status"])
        nnt  = d.get("notes",   row["notes"] or "")
        now  = datetime.now().strftime("%Y-%m-%d %H:%M")
        if session["role"] == "admin" and "is_diplomatic" in d:
            ndip = 1 if d["is_diplomatic"] else 0
            db.execute("UPDATE pilgrims SET status=?,notes=?,is_diplomatic=?,updated_at=? WHERE id=?",
                       (nst,nnt,ndip,now,row["id"]))
        else:
            ndip = row["is_diplomatic"] if "is_diplomatic" in row.keys() else 0
            db.execute("UPDATE pilgrims SET status=?,notes=?,updated_at=? WHERE id=?",(nst,nnt,now,row["id"]))
        db.commit()
        log_action("تعديل", passport,
                   old={"status":old["status"],"notes":old["notes"],"is_diplomatic":old.get("is_diplomatic",0)},
                   new={"status":nst,"notes":nnt,"is_diplomatic":ndip})
        return jsonify({"ok":True})
    finally:
        db.close()

# ── Depart ───────────────────────────────────────────────────

@app.route("/api/depart", methods=["POST"])
@require_login
def depart():
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    db = get_db()
    try:
        row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
        if not row: return jsonify({"ok":False,"msg":"رقم الجواز غير موجود"})
        if session["role"]!="admin" and row["group_id"]!=session["group_id"]:
            return jsonify({"ok":False,"msg":"غير مصرح"})
        if row["departed"]=="نعم":
            return jsonify({"ok":False,"msg":f"غادر مسبقاً في {row['departed_at']}"})
        from datetime import datetime
        batch_id = d.get("batch_id") or "B" + datetime.now().strftime("%Y%m%d%H%M%S")
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        db.execute("UPDATE pilgrims SET departed='نعم',departed_at=?,depart_batch=?,updated_at=? WHERE id=?",(now,batch_id,now,row["id"]))
        db.commit()
        log_action("مغادرة", passport, old={"departed":"لا"}, new={"departed":"نعم","departed_at":now})
        return jsonify({"ok":True,"departed_at":now})
    finally:
        db.close()

# ── List ─────────────────────────────────────────────────────

@app.route("/api/list")
@require_login
def list_records():
    try:
        page = max(1, int(request.args.get("page", 1)))
        per  = min(200, int(request.args.get("per", 50)))
    except (ValueError, TypeError):
        page, per = 1, 50
    gid  = request.args.get("gid","")
    f    = request.args.get("filter","all")
    if session["role"]!="admin":
        gid = session["group_id"]
    if f=="deleted" and session["role"]=="admin":
        cond=["deleted=1"]
    else:
        cond=["deleted=0"]
    params=[]
    if gid: cond.append("group_id=?"); params.append(gid)
    if f=="active":    cond.append("status='مفعل'")
    elif f=="inactive":cond.append("status='غير مفعل'")
    elif f=="departed":cond.append("departed='نعم'")
    elif f=="pending": cond.append("(departed IS NULL OR departed!='نعم')")
    where = "WHERE "+" AND ".join(cond)
    db    = get_db()
    try:
        total = db.execute(f"SELECT COUNT(*) as c FROM pilgrims {where}",params).fetchone()["c"]
        rows  = db.execute(f"SELECT * FROM pilgrims {where} ORDER BY seq LIMIT ? OFFSET ?",
                           params+[per,(page-1)*per]).fetchall()
        return jsonify({"ok":True,"rows":[dict(r) for r in rows],"total":total,"page":page,"per":per})
    finally:
        db.close()

# ── Attachments ──────────────────────────────────────────────

@app.route("/api/attachments")
@require_login
def list_attachments():
    passport = request.args.get("passport","").strip().upper()
    if not passport: return jsonify({"ok":False,"msg":"passport required"}), 400
    db = get_db()
    rows = db.execute(
        "SELECT id, filename, mime, created_at FROM attachments WHERE passport=? ORDER BY id",
        (passport,)
    ).fetchall()
    return jsonify({"ok":True, "attachments":[dict(r) for r in rows]})

@app.route("/api/attachments/<int:att_id>/data")
@require_login
def get_attachment_data(att_id):
    db = get_db()
    row = db.execute("SELECT filename, mime, data FROM attachments WHERE id=?", (att_id,)).fetchone()
    if not row: return jsonify({"ok":False,"msg":"غير موجود"}), 404
    return jsonify({"ok":True, "filename":row["filename"], "mime":row["mime"], "data":row["data"]})

@app.route("/api/attachments/<int:att_id>/raw")
@require_login
def get_attachment_raw(att_id):
    import base64
    db = get_db()
    row = db.execute("SELECT filename, mime, data, filepath FROM attachments WHERE id=?", (att_id,)).fetchone()
    if not row: return ("غير موجود", 404)
    if row["filepath"] and os.path.isfile(row["filepath"]):
        return send_file(row["filepath"], mimetype=row["mime"],
                         download_name=row["filename"], as_attachment=False)
    # fallback: base64 في قاعدة البيانات (بيانات قديمة لم تُنقل بعد)
    if row["data"]:
        raw = base64.b64decode(row["data"])
        return send_file(io.BytesIO(raw), mimetype=row["mime"],
                         download_name=row["filename"], as_attachment=False)
    return ("الملف غير موجود", 404)

@app.route("/api/attachments/add", methods=["POST"])
@require_login
def add_attachment():
    import base64 as b64mod
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    filename = (d.get("filename","") or "").strip()
    mime     = d.get("mime","application/octet-stream") or "application/octet-stream"
    data     = d.get("data","")
    if not passport or not filename or not data:
        return jsonify({"ok":False,"msg":"بيانات ناقصة"}), 400
    db = get_db()
    p = db.execute("SELECT passport FROM pilgrims WHERE passport=? AND deleted=0", (passport,)).fetchone()
    if not p: return jsonify({"ok":False,"msg":"الجواز غير موجود"}), 404
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    # أدرج السجل أولاً للحصول على الـ ID
    cur = db.execute("INSERT INTO attachments (passport,filename,mime,data,created_at) VALUES (?,?,?,?,?)",
                     (passport, filename, mime, '', now))
    att_id = cur.lastrowid
    # احفظ الملف على القرص
    try:
        os.makedirs(UPLOADS_DIR, exist_ok=True)
        _, ext = os.path.splitext(filename)
        ext = re.sub(r'[^a-zA-Z0-9.]', '', ext)[:10]
        fpath = os.path.join(UPLOADS_DIR, f"{att_id}{ext}")
        with open(fpath, 'wb') as f:
            f.write(b64mod.b64decode(data))
        db.execute("UPDATE attachments SET filepath=? WHERE id=?", (fpath, att_id))
    except Exception as e:
        print(f"ERROR saving attachment file: {e}")
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/attachments/all")
@require_admin
def all_attachments():
    db = get_db()
    rows = db.execute(
        "SELECT a.id, a.passport, a.filename, a.mime, a.created_at, p.group_id, g.name as group_name "
        "FROM attachments a "
        "LEFT JOIN pilgrims p ON p.passport=a.passport "
        "LEFT JOIN groups g ON g.id=p.group_id "
        "ORDER BY a.id DESC"
    ).fetchall()
    return jsonify({"ok":True, "attachments":[dict(r) for r in rows]})

@app.route("/api/attachments/delete", methods=["POST"])
@require_login
def delete_attachment():
    d = request.json or {}
    att_id = d.get("id")
    if not att_id: return jsonify({"ok":False,"msg":"id required"}), 400
    db = get_db()
    db.execute("DELETE FROM attachments WHERE id=?", (att_id,))
    db.commit()
    return jsonify({"ok":True})

# ── Delete ───────────────────────────────────────────────────

@app.route("/api/delete", methods=["POST"])
@require_login
def delete():
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    db = get_db()
    try:
        row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
        if not row: return jsonify({"ok":False,"msg":"السجل غير موجود"})
        if session["role"]!="admin" and row["group_id"]!=session["group_id"]:
            return jsonify({"ok":False,"msg":"غير مصرح"})
        old = dict(row)
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        db.execute("UPDATE pilgrims SET deleted=1, updated_at=? WHERE id=?",(now, row["id"]))
        db.commit()
        log_action("حذف", passport, old=old)
        return jsonify({"ok":True})
    finally:
        db.close()

# ── Audit Logs ───────────────────────────────────────────────

@app.route("/api/logs")
@require_admin
def get_logs():
    try:
        page = max(1, int(request.args.get("page", 1)))
        per  = min(200, int(request.args.get("per", 50)))
    except (ValueError, TypeError):
        page, per = 1, 50
    cond=[]; params=[]
    u_f = request.args.get("user","")
    a_f = request.args.get("action","")
    p_f = request.args.get("passport","").upper()
    if u_f: cond.append("username=?"); params.append(u_f)
    if a_f: cond.append("action=?");   params.append(a_f)
    if p_f: cond.append("UPPER(passport) LIKE ?"); params.append(f"%{p_f}%")
    where = ("WHERE "+" AND ".join(cond)) if cond else ""
    db    = get_db()
    try:
        total = db.execute(f"SELECT COUNT(*) as c FROM audit_logs {where}",params).fetchone()["c"]
        rows  = db.execute(f"SELECT * FROM audit_logs {where} ORDER BY id DESC LIMIT ? OFFSET ?",
                           params+[per,(page-1)*per]).fetchall()
        return jsonify({"ok":True,"rows":[dict(r) for r in rows],"total":total})
    finally:
        db.close()

# ── Emp Stats ────────────────────────────────────────────────

@app.route("/api/emp-stats")
@require_admin
def emp_stats():
    period = request.args.get("period","month")
    if   period=="day":   cond="ts LIKE ?"; val=datetime.now().strftime("%Y-%m-%d")+"%"
    elif period=="week":  cond="ts>=?";     val=(datetime.now()-timedelta(days=7)).strftime("%Y-%m-%d")
    else:                 cond="ts LIKE ?"; val=datetime.now().strftime("%Y-%m")+"%"
    db   = get_db()
    try:
        rows = db.execute(f"""SELECT username,user_group,action,COUNT(*) as cnt FROM audit_logs
            WHERE {cond} AND action IN ('إضافة','تعديل','مغادرة','حذف')
            GROUP BY username,user_group,action""",(val,)).fetchall()
        users={}
        for r in rows:
            u=r["username"]
            if u not in users:
                users[u]={"username":u,"group":r["user_group"],"add":0,"edit":0,"depart":0,"delete":0}
            if r["action"]=="إضافة":   users[u]["add"]   +=r["cnt"]
            elif r["action"]=="تعديل": users[u]["edit"]  +=r["cnt"]
            elif r["action"]=="مغادرة":users[u]["depart"]+=r["cnt"]
            elif r["action"]=="حذف":   users[u]["delete"]+=r["cnt"]
        return jsonify({"ok":True,"stats":sorted(users.values(),
                        key=lambda x:x["add"]+x["edit"]+x["depart"],reverse=True)})
    finally:
        db.close()

# ── Users ────────────────────────────────────────────────────

@app.route("/api/users")
@require_admin
def get_users():
    db = get_db()
    try:
        users = db.execute("""SELECT u.id,u.username,u.role,u.group_id,u.active,g.name as group_name
            FROM users u LEFT JOIN groups g ON u.group_id=g.id ORDER BY u.role DESC,u.username""").fetchall()
        return jsonify({"ok":True,"users":[dict(u) for u in users]})
    finally:
        db.close()

@app.route("/api/users/save", methods=["POST"])
@require_admin
def save_user():
    d        = request.json or {}
    uid      = d.get("id")
    username = d.get("username","").strip()
    password = d.get("password","")
    role     = d.get("role","worker")
    group_id = d.get("group_id","") or None
    active   = int(d.get("active",1))
    if role == "worker" and not group_id:
        return jsonify({"ok":False,"msg":"الموظف يجب أن يكون منسوباً لمجموعة"})
    db = get_db()
    try:
        if uid:
            if password:
                db.execute("UPDATE users SET role=?,group_id=?,active=?,password_hash=? WHERE id=?",
                           (role,group_id,active,generate_password_hash(password),uid))
            else:
                db.execute("UPDATE users SET role=?,group_id=?,active=? WHERE id=?",(role,group_id,active,uid))
        else:
            if not username or not password:
                return jsonify({"ok":False,"msg":"اسم المستخدم وكلمة السر مطلوبان"})
            db.execute("INSERT INTO users (username,password_hash,role,group_id,active) VALUES (?,?,?,?,?)",
                       (username,generate_password_hash(password),role,group_id,active))
        db.commit()
        return jsonify({"ok":True})
    except sqlite3.IntegrityError:
        return jsonify({"ok":False,"msg":"اسم المستخدم موجود مسبقاً"})
    finally:
        db.close()

@app.route("/api/users/delete", methods=["POST"])
@require_admin
def delete_user():
    uid = (request.json or {}).get("id")
    db  = get_db()
    try:
        user = db.execute("SELECT * FROM users WHERE id=?",(uid,)).fetchone()
        if not user: return jsonify({"ok":False,"msg":"المستخدم غير موجود"})
        if user["username"] == "admin": return jsonify({"ok":False,"msg":"لا يمكن حذف حساب admin"})
        if str(user["id"]) == str(session.get("user_id")):
            return jsonify({"ok":False,"msg":"لا يمكن حذف حسابك الحالي"})
        db.execute("DELETE FROM users WHERE id=?",(uid,))
        db.commit()
        return jsonify({"ok":True})
    finally:
        db.close()

# ── Groups ───────────────────────────────────────────────────

@app.route("/api/groups")
@require_admin
def get_groups():
    db = get_db()
    try:
        groups = db.execute("SELECT * FROM groups ORDER BY name").fetchall()
        return jsonify({"ok":True,"groups":[dict(g) for g in groups]})
    finally:
        db.close()

@app.route("/api/groups/save", methods=["POST"])
@require_admin
def save_group():
    d      = request.json or {}
    gid    = d.get("id","").strip().lower()
    name   = d.get("name","").strip()
    prefix = d.get("prefix","").strip().upper()
    orig   = d.get("orig_id","")
    if not gid or not name:
        return jsonify({"ok":False,"msg":"المعرف والاسم مطلوبان"})
    if not prefix:
        return jsonify({"ok":False,"msg":"البادئة مطلوبة"})
    db = get_db()
    try:
        dup = db.execute("SELECT id FROM groups WHERE UPPER(prefix)=? AND id!=?",
                         (prefix, orig or gid)).fetchone()
        if dup:
            return jsonify({"ok":False,"msg":f"البادئة '{prefix}' مستخدمة من مجموعة أخرى"})
        icon  = d.get("icon","users").strip() or "users"
        color = d.get("color","#3B82F6").strip() or "#3B82F6"
        if orig and orig != gid:
            db.execute("UPDATE pilgrims SET group_id=? WHERE group_id=?",(gid,orig))
            db.execute("UPDATE users    SET group_id=? WHERE group_id=?",(gid,orig))
            db.execute("DELETE FROM groups WHERE id=?",(orig,))
        db.execute("INSERT OR REPLACE INTO groups (id,name,prefix,icon,color) VALUES (?,?,?,?,?)",
                   (gid,name,prefix,icon,color))
        db.commit()
        log_action("تعديل مجموعة", new={"id":gid,"name":name,"prefix":prefix,"icon":icon,"color":color})
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)})
    finally:
        db.close()

@app.route("/api/groups/delete", methods=["POST"])
@require_admin
def delete_group():
    d          = request.json or {}
    gid        = d.get("id","")
    emp_action = d.get("emp_action","disable")
    db  = get_db()
    try:
        grp = db.execute("SELECT * FROM groups WHERE id=?",(gid,)).fetchone()
        if not grp: return jsonify({"ok":False,"msg":"المجموعة غير موجودة"})
        if emp_action == "delete":
            db.execute("DELETE FROM users WHERE group_id=? AND role='worker'",(gid,))
        else:
            db.execute("UPDATE users SET active=0 WHERE group_id=?",(gid,))
        db.execute("DELETE FROM groups WHERE id=?",(gid,))
        db.commit()
        log_action("حذف مجموعة", old={"id":gid,"name":grp["name"]})
        return jsonify({"ok":True})
    finally:
        db.close()

# ── Shutdown ─────────────────────────────────────────────────

@app.route("/api/shutdown", methods=["POST"])
@require_admin
def shutdown():
    os._exit(0)

RESET_PASSWORD = "NewSeason2025"

@app.route("/api/reset-season", methods=["POST"])
@require_admin
def reset_season():
    data = request.get_json(silent=True) or {}
    if data.get("password") != RESET_PASSWORD:
        return jsonify({"ok": False, "error": "كلمة السر غلط"}), 403
    try:
        do_backup()
    except Exception:
        pass
    db = get_db()
    try:
        db.execute("DELETE FROM pilgrims")
        db.execute("DELETE FROM groups")
        db.execute("DELETE FROM users WHERE role != 'admin'")
        db.execute("DELETE FROM audit_logs")
        db.commit()
    finally:
        db.close()
    return jsonify({"ok": True})

# ── Backup ───────────────────────────────────────────────────

@app.route("/api/backup", methods=["POST"])
@require_admin
def backup():
    try:
        path = do_backup()
        return jsonify({"ok":True,"file":os.path.basename(path)})
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)})

@app.route("/api/backups")
@require_admin
def list_backups():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    files = sorted((f for f in os.listdir(BACKUP_DIR) if f.endswith(".db")), reverse=True)
    return jsonify({"ok":True,"files":files})

# ── Export ───────────────────────────────────────────────────

@app.route("/api/export")
@require_login
def export():
    db      = get_db()
    try:
        req_gid = request.args.get("gid","").strip()
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        ts      = datetime.now().strftime("%Y%m%d_%H%M")

        if session["role"] != "admin":
            grps   = db.execute("SELECT * FROM groups WHERE id=?", (session["group_id"],)).fetchall()
            single = True
        elif req_gid:
            grps   = db.execute("SELECT * FROM groups WHERE id=?", (req_gid,)).fetchall()
            single = True
        else:
            grps   = db.execute("SELECT * FROM groups ORDER BY name").fetchall()
            single = False

        if not grps:
            return jsonify({"ok":False,"msg":"المجموعة غير موجودة"}), 404

        wb  = Workbook(); wb.remove(wb.active)
        NV="1A3A5C"; BL="2563A8"; WH="FFFFFF"; GR="166534"; GRB="DCFCE7"
        RD="991B1B"; RDB="FEE2E2"; OR="9A3412"; ORB="FFEDD5"; AL="F0F5FB"
        def F(c): return PatternFill("solid",fgColor=c)
        def B(c="CBD5E1"):
            s=Side(style='thin',color=c); return Border(left=s,right=s,top=s,bottom=s)
        def fnt(sz=10,b=False,c="1E293B"): return Font(name="Calibri",size=sz,bold=b,color=c)
        def aln(h='center',v='center'): return Alignment(horizontal=h,vertical=v)

        if not single:
            ws=wb.create_sheet("ملخص"); ws.sheet_view.rightToLeft=True
            for col,w in [('A',22),('B',12),('C',12),('D',12),('E',12),('F',14)]:
                ws.column_dimensions[col].width=w
            ws.merge_cells('A1:F1'); c=ws['A1']
            c.value=f"🕋 ملخص الجوازات — {now_str}"
            c.font=Font(name="Calibri",size=14,bold=True,color=WH); c.fill=F(NV); c.alignment=aln()
            ws.row_dimensions[1].height=36
            for i,h in enumerate(["المجموعة","الإجمالي","مفعّل","غير مفعّل","غادر","نسبة المغادرة"],1):
                c=ws.cell(row=2,column=i,value=h)
                c.font=Font(name="Calibri",size=10,bold=True,color=WH); c.fill=F(BL); c.alignment=aln(); c.border=B(WH)
            ws.row_dimensions[2].height=24
            grand=[0,0,0,0]
            for ri,g in enumerate(grps):
                rows=db.execute("SELECT status,departed FROM pilgrims WHERE group_id=? AND deleted=0",(g["id"],)).fetchall()
                t=len(rows); a=sum(1 for r in rows if r["status"]=="مفعل")
                iv=sum(1 for r in rows if r["status"]=="غير مفعل")
                dp=sum(1 for r in rows if r["departed"]=="نعم")
                grand[0]+=t; grand[1]+=a; grand[2]+=iv; grand[3]+=dp
                bg=AL if ri%2==0 else WH; rn=3+ri; ws.row_dimensions[rn].height=22
                ws.cell(rn,1,g["name"]).font=fnt(11,True,NV)
                ws.cell(rn,1).fill=F(bg); ws.cell(rn,1).alignment=aln('right'); ws.cell(rn,1).border=B()
                for ci,val in enumerate([t,a,iv,dp],2):
                    c=ws.cell(rn,ci,val); c.fill=F(bg); c.alignment=aln(); c.border=B()
                    clrs=[(NV,bg),(GR,GRB),(RD,RDB),(OR,ORB)]
                    c.font=Font(name="Calibri",size=12,bold=True,color=clrs[ci-2][0]); c.fill=F(clrs[ci-2][1])
                pct=f"{int(dp/t*100)}%" if t>0 else "—"
                c=ws.cell(rn,6,pct); c.fill=F(bg); c.alignment=aln(); c.border=B(); c.font=fnt(11,True,NV)
            tr=3+len(grps); ws.row_dimensions[tr].height=28
            ws.cell(tr,1,"الإجمالي الكلي").font=Font(name="Calibri",size=11,bold=True,color=WH)
            ws.cell(tr,1).fill=F(NV); ws.cell(tr,1).alignment=aln('right'); ws.cell(tr,1).border=B(NV)
            for ci,val in enumerate(grand,2):
                c=ws.cell(tr,ci,val); c.font=Font(name="Calibri",size=13,bold=True,color=WH)
                c.fill=F(NV); c.alignment=aln(); c.border=B(NV)
            pct=f"{int(grand[3]/grand[0]*100)}%" if grand[0]>0 else "—"
            c=ws.cell(tr,6,pct); c.font=Font(name="Calibri",size=12,bold=True,color=WH)
            c.fill=F(NV); c.alignment=aln(); c.border=B(NV)

            all_rows=db.execute("""
                SELECT p.*, g.name as group_name
                FROM pilgrims p JOIN groups g ON p.group_id=g.id
                WHERE p.deleted=0
                ORDER BY g.name, p.seq
            """).fetchall()
            MHDR=["المجموعة","رقم الجواز","التسلسل","حالة نسك","دبلوماسي","مغادر","وقت المغادرة","ملاحظات","تاريخ الإضافة","آخر تعديل"]
            MCW =[(1,20),(2,22),(3,14),(4,16),(5,12),(6,12),(7,18),(8,32),(9,18),(10,18)]
            wm=wb.create_sheet("كل الجوازات"); wm.sheet_view.rightToLeft=True
            for cw,w in MCW: wm.column_dimensions[chr(64+cw)].width=w
            wm.merge_cells('A1:J1'); c=wm['A1']
            c.value=f"🗂️ جميع الجوازات — {now_str}"
            c.font=Font(name="Calibri",size=14,bold=True,color=WH); c.fill=F(NV); c.alignment=aln()
            wm.row_dimensions[1].height=36
            for ci,h in enumerate(MHDR,1):
                c=wm.cell(2,ci,h); c.font=Font(name="Calibri",size=10,bold=True,color=WH)
                c.fill=F(BL); c.alignment=aln(); c.border=B(WH)
            wm.row_dimensions[2].height=24
            for ri,r in enumerate(all_rows):
                rn=3+ri; bg=AL if ri%2==0 else WH; wm.row_dimensions[rn].height=20
                dipl="نعم" if r["is_diplomatic"] else "لا"
                vals=[r["group_name"], r["passport"],
                      r["seq_code"] or (str(r["seq"]) if r["seq"] else ""),
                      r["status"], dipl, r["departed"],
                      r["departed_at"] or "", r["notes"] or "",
                      r["created_at"] or "", r["updated_at"] or ""]
                for ci,val in enumerate(vals,1):
                    c=wm.cell(rn,ci,val); c.fill=F(bg); c.alignment=aln(); c.border=B()
                    if ci==1:
                        c.font=Font(name="Calibri",size=10,bold=True,color=NV)
                    elif ci==2:
                        c.font=Font(name="Calibri",size=11,bold=True,color=NV)
                        c.alignment=Alignment(horizontal='left',vertical='center')
                    elif ci==3:
                        c.font=Font(name="Calibri",size=11,bold=True,color=BL)
                    elif ci==4:
                        if val=="مفعل": c.fill=F(GRB); c.font=Font(name="Calibri",bold=True,color=GR)
                        elif val=="غير مفعل": c.fill=F(RDB); c.font=Font(name="Calibri",bold=True,color=RD)
                    elif ci==5 and val=="نعم":
                        c.fill=F("F9E8ED"); c.font=Font(name="Calibri",bold=True,color="6D0F2C")
                    elif ci==6 and val=="نعم":
                        c.fill=F(ORB); c.font=Font(name="Calibri",bold=True,color=OR)

        HDR = ["رقم الجواز","التسلسل","حالة نسك","دبلوماسي","مغادر","وقت المغادرة","ملاحظات","آخر تعديل"]
        COL_W = [(1,22),(2,14),(3,16),(4,12),(5,12),(6,18),(7,32),(8,18)]
        for g in grps:
            ws2=wb.create_sheet(g["name"]); ws2.sheet_view.rightToLeft=True
            rows=db.execute("SELECT * FROM pilgrims WHERE group_id=? AND deleted=0 ORDER BY seq",(g["id"],)).fetchall()
            for cw,w in COL_W:
                ws2.column_dimensions[chr(64+cw)].width=w
            ws2.merge_cells('A1:H1'); c=ws2['A1']
            c.value=f"📋 {g['name']} — {now_str}"
            c.font=Font(name="Calibri",size=13,bold=True,color=WH); c.fill=F(NV); c.alignment=aln()
            ws2.row_dimensions[1].height=32
            for ci,h in enumerate(HDR,1):
                c=ws2.cell(2,ci,h); c.font=Font(name="Calibri",size=10,bold=True,color=WH)
                c.fill=F(BL); c.alignment=aln(); c.border=B(WH)
            ws2.row_dimensions[2].height=24
            for ri,r in enumerate(rows):
                rn=3+ri; bg=AL if ri%2==0 else WH; ws2.row_dimensions[rn].height=20
                dipl="نعم" if r["is_diplomatic"] else "لا"
                vals=[r["passport"], r["seq_code"] or (str(r["seq"]) if r["seq"] else ""),
                      r["status"], dipl, r["departed"],
                      r["departed_at"] or "", r["notes"] or "", r["updated_at"] or ""]
                for ci,val in enumerate(vals,1):
                    c=ws2.cell(rn,ci,val); c.fill=F(bg); c.alignment=aln(); c.border=B()
                    if ci==1:
                        c.font=Font(name="Calibri",size=11,bold=True,color=NV)
                        c.alignment=Alignment(horizontal='left',vertical='center')
                    elif ci==2:
                        c.font=Font(name="Calibri",size=11,bold=True,color=BL)
                    elif ci==3:
                        if val=="مفعل": c.fill=F(GRB); c.font=Font(name="Calibri",bold=True,color=GR)
                        elif val=="غير مفعل": c.fill=F(RDB); c.font=Font(name="Calibri",bold=True,color=RD)
                    elif ci==4 and val=="نعم":
                        c.fill=F("F9E8ED"); c.font=Font(name="Calibri",bold=True,color="6D0F2C")
                    elif ci==5 and val=="نعم":
                        c.fill=F(ORB); c.font=Font(name="Calibri",bold=True,color=OR)

        buf=io.BytesIO()
        try:
            wb.save(buf)
        except Exception as e:
            return jsonify({"ok":False,"msg":f"خطأ في بناء الملف: {e}"}), 500
        buf.seek(0)
        if single and grps:
            fname=f"{grps[0]['name']}_{ts}.xlsx"
        else:
            fname=f"جوازات_كامل_{ts}.xlsx"
        return send_file(buf,as_attachment=True,download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        db.close()

@app.route("/api/export-today")
@require_login
def export_today():
    db = get_db()
    try:
        req_gid = request.args.get("gid","").strip()
        today_label = datetime.now().strftime("%Y-%m-%d")
        ts = datetime.now().strftime("%Y%m%d_%H%M")

        if session["role"] != "admin":
            grps = db.execute("SELECT id,name FROM groups WHERE id=?", (session["group_id"],)).fetchall()
            single = True
        elif req_gid:
            grps = db.execute("SELECT id,name FROM groups WHERE id=?", (req_gid,)).fetchall()
            single = True
        else:
            grps = db.execute("SELECT id,name FROM groups ORDER BY name").fetchall()
            single = False

        if not grps:
            return jsonify({"ok":False,"msg":"المجموعة غير موجودة"}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير اليوم"
        ws.sheet_view.rightToLeft = True
        for col, width in [('A',28),('B',14),('C',14),('D',14),('E',16)]:
            ws.column_dimensions[col].width = width

        navy = "1A3A5C"
        blue = "2563A8"
        white = "FFFFFF"
        green_fill = "DCFCE7"
        amber_fill = "FEF3C7"
        slate_fill = "F8FAFC"
        total_fill = "E2E8F0"

        def fill(color): return PatternFill("solid", fgColor=color)
        def border(color="CBD5E1"):
            s = Side(style="thin", color=color)
            return Border(left=s, right=s, top=s, bottom=s)
        def center():
            return Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value = f"طباعة تقرير اليوم — {today_label}"
        c.font = Font(name="Calibri", size=16, bold=True, color=white)
        c.fill = fill(navy)
        c.alignment = center()
        ws.row_dimensions[1].height = 34

        headers = ["اسم المجموعة","عدد الحجاج","عدد المغادرين","عدد المتبقين","نسبة الإنجاز"]
        for idx, header in enumerate(headers, start=1):
            c = ws.cell(2, idx, header)
            c.font = Font(name="Calibri", size=11, bold=True, color=white)
            c.fill = fill(blue)
            c.alignment = center()
            c.border = border(white)
        ws.row_dimensions[2].height = 24

        total_all = departed_all = remaining_all = 0
        for row_idx, g in enumerate(grps, start=3):
            rows = db.execute("SELECT departed FROM pilgrims WHERE group_id=? AND deleted=0", (g["id"],)).fetchall()
            total_count = len(rows)
            departed_count = sum(1 for r in rows if r["departed"]=="نعم")
            remaining_count = total_count - departed_count
            progress = int(departed_count / total_count * 100) if total_count else 0

            total_all += total_count
            departed_all += departed_count
            remaining_all += remaining_count

            base_fill = slate_fill if row_idx % 2 else white
            values = [g["name"], total_count, departed_count, remaining_count, f"{progress}%"]
            for col_idx, val in enumerate(values, start=1):
                c = ws.cell(row_idx, col_idx, val)
                c.border = border()
                c.alignment = center()
                c.fill = fill(base_fill)
                if col_idx == 1:
                    c.font = Font(name="Calibri", size=11, bold=True, color=navy)
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx == 3:
                    c.font = Font(name="Calibri", size=12, bold=True, color="166534")
                    c.fill = fill(green_fill)
                elif col_idx == 4:
                    c.font = Font(name="Calibri", size=12, bold=True, color="92400E")
                    c.fill = fill(amber_fill)
                else:
                    c.font = Font(name="Calibri", size=12, bold=True, color=navy)

        total_row = len(grps) + 3
        total_progress = int(departed_all / total_all * 100) if total_all else 0
        totals = ["الإجمالي", total_all, departed_all, remaining_all, f"{total_progress}%"]
        for col_idx, val in enumerate(totals, start=1):
            c = ws.cell(total_row, col_idx, val)
            c.border = border(navy)
            c.alignment = center()
            c.fill = fill(total_fill if col_idx != 1 else navy)
            c.font = Font(name="Calibri", size=12, bold=True, color=(white if col_idx == 1 else navy))
            if col_idx == 1:
                c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[total_row].height = 26
        ws.freeze_panes = "A3"
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        if single and grps:
            fname = f"{grps[0]['name']}_تقرير_اليوم_{today_label}.xlsx"
        else:
            fname = f"تقرير_اليوم_{today_label}_{ts}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        db.close()

# ── Undelete ─────────────────────────────────────────────────

@app.route("/api/undelete", methods=["POST"])
@require_admin
def undelete():
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    db = get_db()
    try:
        row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
        if not row: return jsonify({"ok":False,"msg":"الجواز غير موجود"})
        if not row["deleted"]: return jsonify({"ok":False,"msg":"السجل غير محذوف"})
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        db.execute("UPDATE pilgrims SET deleted=0, updated_at=? WHERE id=?",(now, row["id"]))
        db.commit()
        log_action("استعادة", passport, old={"deleted":1}, new={"deleted":0})
        return jsonify({"ok":True})
    finally:
        db.close()

# ── Restore & Transfer ───────────────────────────────────────

@app.route("/api/restore-transfer", methods=["POST"])
@require_admin
def restore_transfer():
    d          = request.json or {}
    passport   = d.get("passport","").strip().upper()
    new_gid    = d.get("new_group_id","").strip()
    if not passport or not new_gid:
        return jsonify({"ok":False,"msg":"بيانات ناقصة"})
    db  = get_db()
    try:
        old = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
        if not old:
            return jsonify({"ok":False,"msg":"الجواز غير موجود"})
        if not old["deleted"]:
            return jsonify({"ok":False,"msg":"السجل غير محذوف"})
        if old["group_id"] == new_gid:
            return jsonify({"ok":False,"msg":"الجواز موجود في هذه المجموعة أصلاً، اختر مجموعة مختلفة"})
        new_grp = db.execute("SELECT name FROM groups WHERE id=?",(new_gid,)).fetchone()
        if not new_grp:
            return jsonify({"ok":False,"msg":"المجموعة غير موجودة"})
        old_grp = db.execute("SELECT name FROM groups WHERE id=?",(old["group_id"],)).fetchone()
        old_grp_name = old_grp["name"] if old_grp else old["group_id"]
        seq, seq_code = next_seq_code(db, new_gid)
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        old_seq_str = old['seq_code'] or (str(old['seq']) if old['seq'] else '—')
        transfer_note = f"منقول من {old_grp_name} ({old_seq_str}) إلى {new_grp['name']} ({seq_code})"
        db.execute("BEGIN IMMEDIATE")
        db.execute("""UPDATE pilgrims SET
            group_id=?, seq=?, seq_code=?, status=?,
            is_diplomatic=?, departed='لا', departed_at=NULL,
            notes=?, updated_at=?, deleted=0
            WHERE passport=?""",
            (new_gid, seq, seq_code,
             old["status"] or "غير مفعل",
             old["is_diplomatic"],
             transfer_note, now, passport))
        db.commit()
        log_action("نقل", passport,
                   old={"group":old_grp_name,"seq_code":old["seq_code"]},
                   new={"group":new_grp["name"],"seq_code":seq_code,"note":transfer_note})
        return jsonify({"ok":True,
                        "old_group":old_grp_name,"old_seq":old["seq_code"] or str(old["seq"]),
                        "new_group":new_grp["name"],"new_seq":seq_code})
    finally:
        db.close()

# ── Undepart ─────────────────────────────────────────────────

@app.route("/api/undepart", methods=["POST"])
@require_admin
def undepart():
    d = request.json or {}
    passport = d.get("passport","").strip().upper()
    db = get_db()
    try:
        row = db.execute("SELECT * FROM pilgrims WHERE UPPER(passport)=?",(passport,)).fetchone()
        if not row: return jsonify({"ok":False,"msg":"الجواز غير موجود"})
        if row["departed"] != "نعم":
            return jsonify({"ok":False,"msg":"الجواز لم يُسجَّل مغادراً"})
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        db.execute("UPDATE pilgrims SET departed='لا',departed_at=NULL,updated_at=? WHERE id=?",(now,row["id"]))
        db.commit()
        log_action("إلغاء مغادرة", passport,
                   old={"departed":"نعم","departed_at":row["departed_at"]},
                   new={"departed":"لا"})
        return jsonify({"ok":True})
    finally:
        db.close()

# ── Import ───────────────────────────────────────────────────

@app.route("/api/import", methods=["POST"])
@require_admin
def import_excel():
    from openpyxl import load_workbook as _lw
    if "file" not in request.files:
        return jsonify({"ok":False,"msg":"لم يتم رفع ملف"})
    f   = request.files["file"]
    gid = request.form.get("gid","").strip()
    if not gid:
        return jsonify({"ok":False,"msg":"حدد المجموعة"})
    db  = get_db()
    try:
        if not db.execute("SELECT 1 FROM groups WHERE id=?",(gid,)).fetchone():
            return jsonify({"ok":False,"msg":"المجموعة غير موجودة"})
        try:
            wb = _lw(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
        except Exception:
            return jsonify({"ok":False,"msg":"تعذّر قراءة الملف — تأكد أنه xlsx"})
        skipped=0; errors=[]; valid_rows=[]
        now=datetime.now().strftime("%Y-%m-%d %H:%M")
        for idx,row in enumerate(ws.iter_rows(min_row=1,values_only=True),start=1):
            if not row or not row[0]: continue
            raw=str(row[0]).strip().upper()
            if not re.match(r'^[A-Z]{2}[0-9]{7}$',raw):
                if idx==1: continue
                errors.append({"row":idx,"passport":raw,"reason":"صيغة خاطئة"})
                skipped+=1; continue
            status=str(row[1]).strip() if len(row)>1 and row[1] else ""
            if status not in ("مفعل","غير مفعل"): status="غير مفعل"
            notes=str(row[2]).strip() if len(row)>2 and row[2] else ""
            is_dipl=1 if raw.startswith("DA") else 0
            ex=db.execute("SELECT group_id,deleted FROM pilgrims WHERE passport=?",(raw,)).fetchone()
            if ex and not ex["deleted"]:
                gn=db.execute("SELECT name FROM groups WHERE id=?",(ex["group_id"],)).fetchone()
                errors.append({"row":idx,"passport":raw,
                               "reason":f"مكرر — موجود في {gn['name'] if gn else ex['group_id']}"})
                skipped+=1; continue
            if ex and ex["deleted"]:
                errors.append({"row":idx,"passport":raw,"reason":"كان محذوفاً — يمكن استعادته يدوياً"})
                skipped+=1; continue
            valid_rows.append((raw,status,notes,is_dipl))
        imported=0
        try:
            for raw,status,notes,is_dipl in valid_rows:
                seq,seq_code=next_seq_code(db,gid)
                db.execute("""INSERT INTO pilgrims
                    (passport,group_id,seq,seq_code,status,is_diplomatic,departed,notes,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (raw,gid,seq,seq_code,status,is_dipl,"لا",notes,now,now))
                imported+=1
            db.commit()
        except Exception as e:
            db.rollback()
            return jsonify({"ok":False,"msg":f"فشل الاستيراد — لم يُدرج أي سجل ({e})"})
        log_action("استيراد",new={"group":gid,"imported":imported,"skipped":skipped})
        return jsonify({"ok":True,"imported":imported,"skipped":skipped,"errors":errors[:100]})
    finally:
        db.close()

@app.route('/trips/assign', methods=['POST'])
@require_login
def assign_trip():
    trip_id = request.form.get('trip_id')
    passports_raw = request.form.get('passports')

    if not trip_id:
        return {"ok": False, "error": "missing trip_id"}

    passports = [p.strip() for p in passports_raw.split(',') if p.strip()] if passports_raw else []
    if not passports:
        return {"ok": False, "error": "no passports"}

    placeholders = ','.join(['?'] * len(passports))
    db = get_db()
    try:
        trip = db.execute("SELECT group_id FROM trips WHERE id=?", (trip_id,)).fetchone()
        if not trip:
            return {"ok": False, "error": "trip not found"}
        if session.get("role") != "admin" and trip["group_id"] != session.get("group_id"):
            return {"ok": False, "error": "unauthorized"}
        previous_assignments = db.execute(
            f"""SELECT passport, trip_id
                FROM pilgrims
                WHERE passport IN ({placeholders})
                  AND deleted=0
                  AND (departed IS NULL OR departed!='نعم')
                  AND group_id=?""",
            passports + [trip["group_id"]]
        ).fetchall()
        db.execute(
            f"""UPDATE pilgrims
                SET trip_id=?
                WHERE passport IN ({placeholders})
                  AND deleted=0
                  AND (departed IS NULL OR departed!='نعم')
                  AND group_id=?""",
            [trip_id] + passports + [trip["group_id"]]
        )
        db.commit()
        for pilgrim in previous_assignments:
            if pilgrim["trip_id"] != trip_id:
                log_action(
                    'trip_assign',
                    pilgrim["passport"],
                    {'trip_id': pilgrim["trip_id"]},
                    {'trip_id': trip_id}
                )
        return {"ok": True, "updated": len(passports)}
    finally:
        db.close()

@app.route('/api/pilgrims/unassigned')
@require_login
def unassigned_pilgrims():
    db = get_db()
    try:
        where = [
            "p.deleted=0",
            "p.status='مفعل'",
            "(p.departed IS NULL OR p.departed!='نعم')",
            "(p.trip_id IS NULL OR TRIM(p.trip_id)='')",
        ]
        params = []
        if session.get("role") != "admin":
            where.append("p.group_id=?")
            params.append(session.get("group_id"))
        rows = db.execute(f"""
            SELECT p.passport, p.seq_code, p.status, p.group_id, g.name AS group_name
            FROM pilgrims p
            LEFT JOIN groups g ON g.id=p.group_id
            WHERE {" AND ".join(where)}
            ORDER BY g.name, p.seq, p.passport
        """, params).fetchall()
        pilgrims = [dict(row) for row in rows]
        return jsonify({"ok": True, "count": len(pilgrims), "pilgrims": pilgrims})
    finally:
        db.close()

@app.route('/trips/create', methods=['POST'])
@require_login
def create_trip():
    trip_id = request.form.get('id')
    if not trip_id:
        return {"ok": False, "error": "missing id"}

    if session.get("role") == "admin":
        group_id = request.form.get('group_id') or session.get("group_id")
    else:
        group_id = session.get("group_id")
    hotel_name = request.form.get('hotel_name')
    location_url = request.form.get('location_url')
    nationality = request.form.get('nationality')
    flight_no = request.form.get('flight_no')
    carrier = request.form.get('carrier')
    destination = request.form.get('destination')
    approval_no = request.form.get('approval_no')
    departure_time = request.form.get('departure_time')
    housing_contract_no = request.form.get('housing_contract_no')
    bus_stand_time = request.form.get('bus_stand_time')
    bus_departure_time = request.form.get('bus_departure_time')
    bus_count = request.form.get('bus_count')
    planned_pilgrim_count = request.form.get('planned_pilgrim_count')
    report_date = request.form.get('report_date') or datetime.now().strftime("%Y-%m-%d")
    notes = request.form.get('notes')
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    if not group_id:
        return {"ok": False, "error": "missing group_id"}

    db = get_db()
    try:
        if not db.execute("SELECT 1 FROM groups WHERE id=?", (group_id,)).fetchone():
            return {"ok": False, "error": "group not found"}
        db.execute("""
            INSERT INTO trips (
                id, group_id, hotel_name, location_url, nationality,
                flight_no, carrier, destination, approval_no,
                departure_time, housing_contract_no, bus_stand_time,
                bus_departure_time, bus_count, planned_pilgrim_count, report_date, notes, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            trip_id, group_id, hotel_name, location_url, nationality,
            flight_no, carrier, destination, approval_no,
            departure_time, housing_contract_no, bus_stand_time,
            bus_departure_time, bus_count, planned_pilgrim_count, report_date, notes, created_at
        ))
        db.commit()
        return {"ok": True}
    except sqlite3.IntegrityError:
        return {"ok": False, "error": "trip exists"}
    finally:
        db.close()

@app.route('/trips/update', methods=['POST'])
@require_login
def update_trip():
    trip_id = (request.form.get('trip_id') or request.form.get('id') or '').strip()
    if not trip_id:
        return {"ok": False, "error": "missing trip_id"}

    trip_fields = [
        'report_date',
        'hotel_name',
        'location_url',
        'nationality',
        'flight_no',
        'carrier',
        'destination',
        'approval_no',
        'departure_time',
        'housing_contract_no',
        'bus_stand_time',
        'bus_departure_time',
        'planned_pilgrim_count',
        'bus_count',
        'notes',
    ]

    db = get_db()
    try:
        trip = db.execute("SELECT * FROM trips WHERE id=?", (trip_id,)).fetchone()
        if not trip:
            return {"ok": False, "error": "trip not found"}, 404
        if session.get("role") != "admin" and trip["group_id"] != session.get("group_id"):
            return {"ok": False, "error": "unauthorized"}, 403

        if session.get("role") == "admin":
            group_id = (request.form.get('group_id') or trip["group_id"] or '').strip()
        else:
            group_id = session.get("group_id")

        if not group_id:
            return {"ok": False, "error": "missing group_id"}, 400
        if not db.execute("SELECT 1 FROM groups WHERE id=?", (group_id,)).fetchone():
            return {"ok": False, "error": "group not found"}, 400

        if group_id != trip["group_id"]:
            foreign_linked = db.execute("""
                SELECT COUNT(*) AS cnt
                FROM pilgrims
                WHERE trip_id=?
                  AND deleted=0
                  AND group_id!=?
            """, (trip_id, group_id)).fetchone()["cnt"]
            if foreign_linked:
                return {"ok": False, "error": "linked pilgrims belong to another group"}, 400

        updates = {"group_id": group_id}
        for field in trip_fields:
            updates[field] = request.form.get(field)
        if not updates["report_date"]:
            updates["report_date"] = trip["report_date"] or datetime.now().strftime("%Y-%m-%d")

        set_clause = ", ".join(f"{field}=?" for field in updates.keys())
        db.execute(f"UPDATE trips SET {set_clause} WHERE id=?", list(updates.values()) + [trip_id])
        db.commit()
        log_action(
            "trip_update",
            trip_id,
            old={"group_id": trip["group_id"], "report_date": trip["report_date"]},
            new={"group_id": group_id, "report_date": updates["report_date"]},
        )
        return {"ok": True}
    finally:
        db.close()

@app.route('/trips/list')
@require_login
def list_trips():
    report_date = request.args.get("date", "").strip()
    db = get_db()
    try:
        if session.get("role") == "admin":
            where = "WHERE t.report_date=?" if report_date else ""
            params = (report_date,) if report_date else ()
            rows = db.execute(f"""
                SELECT t.*, g.name AS group_name
                FROM trips t LEFT JOIN groups g ON g.id=t.group_id
                {where}
                ORDER BY t.created_at DESC
            """, params).fetchall()
        else:
            date_filter = "AND t.report_date=?" if report_date else ""
            params = [session.get("group_id")]
            if report_date:
                params.append(report_date)
            rows = db.execute(f"""
                SELECT t.*, g.name AS group_name
                FROM trips t LEFT JOIN groups g ON g.id=t.group_id
                WHERE t.group_id=?
                {date_filter}
                ORDER BY t.created_at DESC
            """, params).fetchall()
        return {"ok": True, "trips": [dict(row) for row in rows]}
    finally:
        db.close()

@app.route('/trips/depart', methods=['POST'])
@require_login
def depart_trip():
    trip_id = request.form.get('trip_id')
    if not trip_id:
        return {"ok": False, "error": "missing trip_id"}

    batch_id = trip_id
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    db = get_db()
    try:
        trip = db.execute("SELECT group_id, status FROM trips WHERE id=?", (trip_id,)).fetchone()
        if not trip:
            return {"ok": False, "error": "trip not found"}
        if session.get("role") != "admin" and trip["group_id"] != session.get("group_id"):
            return {"ok": False, "error": "unauthorized"}
        if trip['status'] != 'confirmed':
            return {"ok": False, "error": "يجب تأكيد الرحلة قبل تسجيل المغادرة"}, 400
        cur = db.execute("""
            UPDATE pilgrims
            SET departed='نعم',
                departed_at=?,
                depart_batch=?,
                updated_at=?
            WHERE trip_id=?
              AND deleted=0
        """, (now, batch_id, now, trip_id))
        db.execute("UPDATE trips SET status='departed' WHERE id=?", (trip_id,))
        db.commit()
        return {"ok": True, "updated": cur.rowcount}
    finally:
        db.close()

@app.route('/trips/undepart', methods=['POST'])
@require_admin
def undepart_trip():
    data = request.get_json(silent=True) or {}
    trip_id = (data.get('trip_id') or '').strip()
    if not trip_id:
        return jsonify({"ok": False, "error": "missing trip_id"}), 400

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    db = get_db()
    try:
        trip = db.execute("SELECT id, status FROM trips WHERE id=?", (trip_id,)).fetchone()
        if not trip:
            return jsonify({"ok": False, "error": "trip not found"}), 404
        if trip["status"] != "departed":
            return jsonify({"ok": False, "error": "trip is not departed"}), 400

        departed_count = db.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM pilgrims
            WHERE trip_id=?
              AND deleted=0
              AND (departed_at IS NOT NULL OR depart_batch IS NOT NULL)
            """,
            (trip_id,),
        ).fetchone()["cnt"]

        db.execute(
            """
            UPDATE pilgrims
            SET departed='لا',
                departed_at=NULL,
                depart_batch=NULL,
                updated_at=?
            WHERE trip_id=?
              AND deleted=0
              AND (departed_at IS NOT NULL OR depart_batch IS NOT NULL)
            """,
            (now, trip_id),
        )
        db.execute("UPDATE trips SET status=? WHERE id=?", ("confirmed", trip_id))
        db.commit()
        log_action(
            "trip_undepart",
            trip_id,
            old={"status": trip["status"], "departed_count": departed_count},
            new={"status": "confirmed", "departed_count": 0},
        )
        return jsonify({"ok": True, "updated": departed_count, "status": "confirmed"})
    finally:
        db.close()

@app.route('/trips/update-status', methods=['POST'])
@require_login
def update_trip_status():
    data = request.get_json(force=True)
    trip_id = (data.get('trip_id') or '').strip()
    new_status = (data.get('status') or '').strip()
    if not trip_id or new_status != 'confirmed':
        return {"ok": False, "error": "بيانات غير صالحة"}, 400
    db = get_db()
    try:
        trip = db.execute("SELECT group_id, status FROM trips WHERE id=?", (trip_id,)).fetchone()
        if not trip:
            return {"ok": False, "error": "الرحلة غير موجودة"}, 404
        if session.get('role') != 'admin' and trip['group_id'] != session.get('group_id'):
            return {"ok": False, "error": "غير مصرح"}, 403
        if trip['status'] != 'draft':
            return {"ok": False, "error": "لا يمكن تأكيد هذه الحالة"}, 400
        db.execute("UPDATE trips SET status=? WHERE id=?", (new_status, trip_id))
        db.commit()
        log_action('trip_status', trip_id, {'status': trip['status']}, {'status': new_status})
    finally:
        db.close()
    return {"ok": True}

@app.route('/trips/details')
@require_login
def trip_details():
    trip_id = request.args.get('trip_id', '').strip()
    if not trip_id:
        return {"ok": False, "error": "trip_id مطلوب"}, 400
    db = get_db()
    try:
        trip = db.execute("SELECT * FROM trips WHERE id=?", (trip_id,)).fetchone()
        if not trip:
            return {"ok": False, "error": "الرحلة غير موجودة"}, 404
        if session.get('role') != 'admin' and trip['group_id'] != session.get('group_id'):
            return {"ok": False, "error": "غير مصرح"}, 403
        pilgrims = db.execute(
            "SELECT passport, seq_code, status, departed, notes FROM pilgrims "
            "WHERE trip_id=? AND deleted=0",
            (trip_id,)
        ).fetchall()
    finally:
        db.close()
    return {
        "ok": True,
        "trip": dict(trip),
        "pilgrims": [dict(p) for p in pilgrims],
        "actual_count": len(pilgrims)
    }

def _docx_text(text):
    return escape(str(text or ""))

def _docx_cell(text, shade=None, bold=False):
    fill = f'<w:shd w:fill="{shade}"/>' if shade else ""
    b = "<w:b/>" if bold else ""
    return (
        "<w:tc><w:tcPr>"
        '<w:tcW w:w="1700" w:type="dxa"/>'
        f"{fill}</w:tcPr>"
        '<w:p><w:pPr><w:bidi/><w:jc w:val="center"/></w:pPr>'
        f'<w:r><w:rPr><w:rtl/>{b}<w:sz w:val="20"/></w:rPr><w:t>{_docx_text(text)}</w:t></w:r>'
        "</w:p></w:tc>"
    )

def _docx_paragraph(text, size=28, bold=False, center=True):
    b = "<w:b/>" if bold else ""
    jc = '<w:jc w:val="center"/>' if center else '<w:jc w:val="right"/>'
    return (
        f'<w:p><w:pPr><w:bidi/>{jc}</w:pPr>'
        f'<w:r><w:rPr><w:rtl/>{b}<w:sz w:val="{size}"/></w:rPr><w:t>{_docx_text(text)}</w:t></w:r>'
        "</w:p>"
    )

def _arabic_report_date(report_date):
    try:
        dt = datetime.strptime(report_date, "%Y-%m-%d")
        weekdays = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
        return f"{weekdays[dt.weekday()]} الموافق /{dt.strftime('%d/%m/%Y')}"
    except ValueError:
        return f"الموافق /{report_date}"

def _arabic_report_date_safe(report_date):
    try:
        dt = datetime.strptime(report_date, "%Y-%m-%d")
        weekdays = [
            "\u0627\u0644\u0627\u062b\u0646\u064a\u0646",
            "\u0627\u0644\u062b\u0644\u0627\u062b\u0627\u0621",
            "\u0627\u0644\u0623\u0631\u0628\u0639\u0627\u0621",
            "\u0627\u0644\u062e\u0645\u064a\u0633",
            "\u0627\u0644\u062c\u0645\u0639\u0629",
            "\u0627\u0644\u0633\u0628\u062a",
            "\u0627\u0644\u0623\u062d\u062f",
        ]
        return f"{weekdays[dt.weekday()]} \u0627\u0644\u0645\u0648\u0627\u0641\u0642 /{dt.strftime('%d/%m/%Y')}"
    except ValueError:
        return f"\u0627\u0644\u0645\u0648\u0627\u0641\u0642 /{report_date}"

def _set_docx_text(container, value):
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    texts = container.findall(".//w:t", ns)
    if not texts:
        return
    texts[0].text = "" if value is None else str(value)
    for text in texts[1:]:
        text.text = ""

def _trip_docx_row_values(row, idx):
    return [
        row["notes"] or "-",
        row["bus_departure_time"] or "",
        row["bus_stand_time"] or "",
        row["housing_contract_no"] or "",
        row["departure_time"] or "",
        row["flight_no"] or "",
        row["pilgrims_count"] or 0,
        row["bus_count"] or "",
        row["approval_no"] or "-",
        row["destination"] or "",
        row["carrier"] or "",
        row["nationality"] or "-",
        row["location_url"] or "-",
        row["hotel_name"] or "",
        idx,
    ]

def _fill_trip_template_page(page_elements, group_name, rows, report_date, ns):
    title = (
        "\u062c\u062f\u0648\u0644 \u0627\u0644\u0631\u062d\u0644\u0627\u062a "
        "\u0627\u0644\u0645\u063a\u0627\u062f\u0631\u0629 \u0644\u064a\u0640\u0648\u0645/"
        f"{_arabic_report_date_safe(report_date)} - {group_name}"
    )
    title_marker = "\u062c\u062f\u0648\u0644 \u0627\u0644\u0631\u062d\u0644\u0627\u062a"
    for element in page_elements:
        paragraphs = []
        if element.tag == f"{{{ns['w']}}}p":
            paragraphs.append(element)
        paragraphs.extend(element.findall(".//w:p", ns))
        for paragraph in paragraphs:
            paragraph_text = "".join(t.text or "" for t in paragraph.findall(".//w:t", ns))
            if title_marker in paragraph_text:
                _set_docx_text(paragraph, title)
                break

    table = None
    for element in page_elements:
        if element.tag == f"{{{ns['w']}}}tbl":
            table = element
            break
        table = element.find(".//w:tbl", ns)
        if table is not None:
            break
    if table is None:
        return False

    table_rows = table.findall("w:tr", ns)
    if len(table_rows) < 2:
        return False

    data_template = deepcopy(table_rows[1])
    for old_row in table_rows[1:]:
        table.remove(old_row)

    for idx, row in enumerate(rows, start=1):
        new_row = deepcopy(data_template)
        for cell, value in zip(new_row.findall("w:tc", ns), _trip_docx_row_values(row, idx)):
            _set_docx_text(cell, value)
        table.append(new_row)

    for _ in range(max(0, 8 - len(rows))):
        empty_row = deepcopy(data_template)
        for cell in empty_row.findall("w:tc", ns):
            _set_docx_text(cell, "")
        table.append(empty_row)

    return True

def _docx_page_break(ns):
    paragraph = ET.Element(f"{{{ns['w']}}}p")
    run = ET.SubElement(paragraph, f"{{{ns['w']}}}r")
    br = ET.SubElement(run, f"{{{ns['w']}}}br")
    br.set(f"{{{ns['w']}}}type", "page")
    return paragraph

def _build_daily_trips_docx_from_template(groups_rows, report_date):
    if not os.path.exists(TRIPS_DAILY_TEMPLATE):
        return None

    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    ET.register_namespace("w", ns["w"])
    ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")
    ET.register_namespace("wp", "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing")
    ET.register_namespace("a", "http://schemas.openxmlformats.org/drawingml/2006/main")
    ET.register_namespace("pic", "http://schemas.openxmlformats.org/drawingml/2006/picture")

    with zipfile.ZipFile(TRIPS_DAILY_TEMPLATE, "r") as template:
        document_xml = template.read("word/document.xml")

    root = ET.fromstring(document_xml)
    title = f"جدول الرحلات المغادرة ليـوم/{_arabic_report_date(report_date)}"
    for paragraph in root.findall(".//w:p", ns):
        paragraph_text = "".join(t.text or "" for t in paragraph.findall(".//w:t", ns))
        if "جدول الرحلات" in paragraph_text:
            _set_docx_text(paragraph, title)
            break

    table = root.find(".//w:tbl", ns)
    if table is None:
        return None

    table_rows = table.findall("w:tr", ns)
    if len(table_rows) < 2:
        return None

    data_template = deepcopy(table_rows[1])
    for old_row in table_rows[1:]:
        table.remove(old_row)

    flat_rows = []
    for _group_name, rows in groups_rows:
        flat_rows.extend(rows)

    for idx, row in enumerate(flat_rows, start=1):
        new_row = deepcopy(data_template)
        for cell, value in zip(new_row.findall("w:tc", ns), _trip_docx_row_values(row, idx)):
            _set_docx_text(cell, value)
        table.append(new_row)

    for _ in range(max(0, 8 - len(flat_rows))):
        empty_row = deepcopy(data_template)
        for cell in empty_row.findall("w:tc", ns):
            _set_docx_text(cell, "")
        table.append(empty_row)

    buf = io.BytesIO()
    with zipfile.ZipFile(TRIPS_DAILY_TEMPLATE, "r") as template:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as output:
            for item in template.infolist():
                if item.filename == "word/document.xml":
                    output.writestr(item, ET.tostring(root, encoding="utf-8", xml_declaration=True))
                else:
                    output.writestr(item, template.read(item.filename))
    buf.seek(0)
    return buf

def _build_daily_trips_docx_group_pages(groups_rows, report_date):
    if not os.path.exists(TRIPS_DAILY_TEMPLATE):
        return None

    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    ET.register_namespace("w", ns["w"])
    ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")
    ET.register_namespace("wp", "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing")
    ET.register_namespace("a", "http://schemas.openxmlformats.org/drawingml/2006/main")
    ET.register_namespace("pic", "http://schemas.openxmlformats.org/drawingml/2006/picture")

    with zipfile.ZipFile(TRIPS_DAILY_TEMPLATE, "r") as template:
        root = ET.fromstring(template.read("word/document.xml"))

    body = root.find("w:body", ns)
    if body is None:
        return None

    original_children = list(body)
    sect_pr = next((child for child in original_children if child.tag == f"{{{ns['w']}}}sectPr"), None)
    page_template = [deepcopy(child) for child in original_children if child is not sect_pr]
    if not page_template:
        return None

    for child in list(body):
        body.remove(child)

    if not groups_rows:
        groups_rows = [("", [])]

    for page_idx, (group_name, rows) in enumerate(groups_rows):
        page_elements = deepcopy(page_template)
        if not _fill_trip_template_page(page_elements, group_name, rows, report_date, ns):
            return None
        for element in page_elements:
            body.append(element)
        if page_idx < len(groups_rows) - 1:
            body.append(_docx_page_break(ns))

    if sect_pr is not None:
        body.append(deepcopy(sect_pr))

    buf = io.BytesIO()
    with zipfile.ZipFile(TRIPS_DAILY_TEMPLATE, "r") as template:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as output:
            for item in template.infolist():
                if item.filename == "word/document.xml":
                    output.writestr(item, ET.tostring(root, encoding="utf-8", xml_declaration=True))
                else:
                    output.writestr(item, template.read(item.filename))
    buf.seek(0)
    return buf

def _build_daily_trips_docx(groups_rows, report_date):
    templated = _build_daily_trips_docx_group_pages(groups_rows, report_date)
    if templated:
        return templated

    headers = [
        "الملاحظات", "وقت مغادرة الحافلة", "وقت وقوف الحافلة", "رقم عقد السكن",
        "موعد الاقلاع", "رقم الرحلة", "اجمالي عدد الحجاج", "عدد الحافلات",
        "رقم الاعتماد", "الوجهة", "الشركة الناقلة", "الجنسية", "رابط الموقع",
        "اسم الفندق", "م"
    ]
    body = [
        _docx_paragraph("مركز اثراء الجود (  1   )", size=24, bold=True),
        _docx_paragraph(f"جدول الرحلات المغادرة ليوم / {report_date} الموافق / {report_date}", size=30, bold=True),
    ]
    for group_name, rows in groups_rows:
        body.append(_docx_paragraph(group_name, size=24, bold=True, center=False))
        table = [
            '<w:tbl><w:tblPr><w:bidiVisual/><w:tblW w:w="0" w:type="auto"/>'
            '<w:tblBorders><w:top w:val="single" w:sz="6"/><w:left w:val="single" w:sz="6"/>'
            '<w:bottom w:val="single" w:sz="6"/><w:right w:val="single" w:sz="6"/>'
            '<w:insideH w:val="single" w:sz="6"/><w:insideV w:val="single" w:sz="6"/></w:tblBorders>'
            "</w:tblPr>"
        ]
        table.append("<w:tr>" + "".join(_docx_cell(h, "D9EAF7", True) for h in headers) + "</w:tr>")
        for idx, row in enumerate(rows, start=1):
            values = [
                row["notes"] or "-",
                row["bus_departure_time"] or "",
                row["bus_stand_time"] or "",
                row["housing_contract_no"] or "",
                row["departure_time"] or "",
                row["flight_no"] or "",
                row["pilgrims_count"] or 0,
                row["bus_count"] or "",
                row["approval_no"] or "-",
                row["destination"] or "",
                row["carrier"] or "",
                row["nationality"] or "-",
                row["location_url"] or "-",
                row["hotel_name"] or "",
                idx,
            ]
            table.append("<w:tr>" + "".join(_docx_cell(v) for v in values) + "</w:tr>")
        if not rows:
            empty = [""] * (len(headers) - 1) + ["-"]
            table.append("<w:tr>" + "".join(_docx_cell(v) for v in empty) + "</w:tr>")
        table.append("</w:tbl>")
        body.append("".join(table))
        body.append(_docx_paragraph("ختم المركز: ____________", size=24, bold=True, center=False))

    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:body>'
        + "".join(body) +
        '<w:sectPr><w:pgSz w:w="16838" w:h="11906" w:orient="landscape"/>'
        '<w:pgMar w:top="720" w:right="420" w:bottom="720" w:left="420" w:header="720" w:footer="720" w:gutter="0"/>'
        '</w:sectPr></w:body></w:document>'
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as docx:
        docx.writestr("[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            '</Types>')
        docx.writestr("_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            '</Relationships>')
        docx.writestr("word/document.xml", document_xml)
    buf.seek(0)
    return buf

@app.route('/trips/daily-report')
@require_login
def daily_trips_report():
    report_date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")
    req_gid = request.args.get("gid", "").strip()
    db = get_db()
    try:
        if session.get("role") == "admin":
            if req_gid:
                groups = db.execute("SELECT id, name FROM groups WHERE id=?", (req_gid,)).fetchall()
            else:
                groups = db.execute("SELECT id, name FROM groups ORDER BY name").fetchall()
        else:
            groups = db.execute("SELECT id, name FROM groups WHERE id=?", (session.get("group_id"),)).fetchall()

        groups_rows = []
        for group in groups:
            rows = db.execute("""
                SELECT t.id, t.hotel_name, t.location_url, t.nationality,
                       t.flight_no, t.carrier, t.destination, t.approval_no,
                       t.departure_time, t.housing_contract_no, t.bus_stand_time,
                       t.bus_departure_time, t.bus_count, t.notes,
                       CASE
                           WHEN t.planned_pilgrim_count IS NOT NULL THEN t.planned_pilgrim_count
                           ELSE COUNT(p.id)
                       END AS pilgrims_count
                FROM trips t
                LEFT JOIN pilgrims p ON p.trip_id=t.id AND p.deleted=0
                WHERE t.group_id=?
                  AND t.report_date=?
                GROUP BY t.id
                ORDER BY t.bus_departure_time, t.departure_time, t.id
            """, (group["id"], report_date)).fetchall()
            groups_rows.append((group["name"], rows))
    finally:
        db.close()

    buf = _build_daily_trips_docx(groups_rows, report_date)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"daily_trips_{report_date}.docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )

@app.route('/trips/departed-report')
@require_login
def departed_trips_report():
    req_gid = request.args.get("gid", "").strip()

    db = get_db()
    try:
        where = [
            "p.departed='نعم'",
            "p.deleted=0",
            "p.trip_id IS NOT NULL",
        ]
        params = []
        if session.get("role") == "admin":
            if req_gid:
                where.append("p.group_id=?")
                params.append(req_gid)
        else:
            current_group = session.get("group_id")
            if not current_group:
                return {"ok": False, "error": "missing group_id"}
            where.append("p.group_id=?")
            params.append(current_group)

        rows = db.execute(f"""
            SELECT p.trip_id,
                   t.hotel_name,
                   t.flight_no,
                   t.carrier,
                   t.destination,
                   t.departure_time,
                   t.bus_count,
                   COUNT(*) AS pilgrims_count,
                   MIN(p.departed_at) AS departed_time
            FROM pilgrims p
            JOIN trips t ON p.trip_id = t.id
            WHERE {" AND ".join(where)}
            GROUP BY p.trip_id
            ORDER BY MIN(p.departed_at)
        """, params).fetchall()
    finally:
        db.close()

    today_label = datetime.now().strftime("%Y-%m-%d")
    headers = [
        "وقت المغادرة", "اجمالي عدد الحجاج", "رقم الرحلة",
        "الوجهة", "الشركة الناقلة", "اسم الفندق", "م"
    ]
    body = [
        _docx_paragraph("مركز اثراء الجود", size=24, bold=True),
        _docx_paragraph("جدول الرحلات المغادرة", size=30, bold=True),
        _docx_paragraph(f"ليوم / {today_label} الموافق / {today_label}", size=24, bold=True),
    ]
    tbl = [
        '<w:tbl><w:tblPr><w:bidiVisual/><w:tblW w:w="0" w:type="auto"/>'
        '<w:tblBorders><w:top w:val="single" w:sz="6"/><w:left w:val="single" w:sz="6"/>'
        '<w:bottom w:val="single" w:sz="6"/><w:right w:val="single" w:sz="6"/>'
        '<w:insideH w:val="single" w:sz="6"/><w:insideV w:val="single" w:sz="6"/></w:tblBorders>'
        '</w:tblPr>'
    ]
    tbl.append("<w:tr>" + "".join(_docx_cell(h, "D9EAF7", True) for h in headers) + "</w:tr>")
    for idx, row in enumerate(rows, start=1):
        values = [
            row["departed_time"] or "",
            row["pilgrims_count"],
            row["flight_no"] or "",
            row["destination"] or "",
            row["carrier"] or "",
            row["hotel_name"] or "",
            idx,
        ]
        tbl.append("<w:tr>" + "".join(_docx_cell(v) for v in values) + "</w:tr>")
    if not rows:
        empty = [""] * (len(headers) - 1) + ["-"]
        tbl.append("<w:tr>" + "".join(_docx_cell(v) for v in empty) + "</w:tr>")
    tbl.append("</w:tbl>")
    body.append("".join(tbl))
    body.append(_docx_paragraph("ختم المركز: ____________", size=24, bold=True, center=False))

    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:body>'
        + "".join(body) +
        '<w:sectPr><w:pgSz w:w="16838" w:h="11906" w:orient="landscape"/>'
        '<w:pgMar w:top="720" w:right="420" w:bottom="720" w:left="420" w:header="720" w:footer="720" w:gutter="0"/>'
        '</w:sectPr></w:body></w:document>'
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as docx:
        docx.writestr("[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            '</Types>')
        docx.writestr("_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            '</Relationships>')
        docx.writestr("word/document.xml", document_xml)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="departed_trips_report.docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )

# ── HTML ─────────────────────────────────────────────────────

@app.route("/scan")
def scan_page():
    if not session.get("user_id"):
        return redirect("/")
    return render_template("scan.html")

@app.route("/scan-depart")
def scan_depart_page():
    if not session.get("user_id"):
        return redirect("/")
    return redirect("/scan?mode=depart")

@app.route("/dashboard")
def dashboard_page():
    if not session.get("user_id"):
        return redirect("/")
    return render_template("dashboard.html")

@app.route("/")
def index():
    return render_template("index.html")




if __name__ == "__main__":
    init_db()
    migrate_json()
    migrate_seq_codes()
    os.makedirs(BACKUP_DIR, exist_ok=True)
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    migrate_attachments_to_fs()
    threading.Thread(target=backup_scheduler, daemon=True).start()

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80)); ip = s.getsockname()[0]
    except:
        ip = "127.0.0.1"
    finally:
        s.close()

    print("=" * 52)
    print("🕋  نظام الجوازات — v2.0")
    print("=" * 52)
    print(f"\n✅ السيرفر شغّال!")
    print(f"\n🔗 جهازك:    http://localhost:5000")
    print(f"🔗 الشبكة:   http://{ip}:5000")
    print("\n📋 بيانات الدخول:")
    print("   مشرف:  admin / admin2024")
    print("   موظف:  fagh  / 1234  (مثال)")
    print("\n⚠️  لا تغلق هذه النافذة أثناء العمل")
    print("=" * 52)
    app.run(host="0.0.0.0", port=5000, debug=False)
